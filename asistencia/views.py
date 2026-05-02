from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.utils.timezone import localdate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Asistencia
from .forms import AsistenciaFiltroForm
from academico.models import CargaAcademica, Estudiante, Grado, Grupo, PeriodoAcademico
from alertas.models import ConfiguracionAlerta
from alertas.utils import evaluar_alertas_academicas, comparar_valor
from evaluacion.models import ActividadEvaluativa, Calificacion
from usuarios.auditoria import registrar_auditoria_cambio
from usuarios.decorators import role_required
from usuarios.notificaciones import crear_notificaciones_docentes
from usuarios.permissions import cargas_visibles_para, filtrar_estudiantes_visibles, puede_gestionar_docencia, no_es_estudiante


def _preparar_hoja_excel(hoja, titulo, encabezados):
    hoja.title = titulo
    hoja.append(encabezados)
    fill = PatternFill('solid', fgColor='1F2937')
    for celda in hoja[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = fill


def _ajustar_columnas(hoja):
    for columna in hoja.columns:
        longitud = max(len(str(celda.value or '')) for celda in columna)
        hoja.column_dimensions[get_column_letter(columna[0].column)].width = min(longitud + 2, 45)


def _grados_y_grupos_desde_cargas(cargas_queryset):
    grupos = Grupo.objects.select_related('grado').filter(
        id__in=cargas_queryset.values_list('grupo_id', flat=True).distinct()
    ).order_by('grado__nombre', 'nombre')
    grados = Grado.objects.filter(
        id__in=grupos.values_list('grado_id', flat=True).distinct()
    ).order_by('nombre')
    return grados, grupos


def _datos_resumen_asistencia(request):
    grado_id = request.GET.get('grado', '').strip()
    grupo_id = request.GET.get('grupo', '').strip()
    busqueda = request.GET.get('busqueda', '').strip()

    cargas_visibles = cargas_visibles_para(request.user)
    grados, grupos = _grados_y_grupos_desde_cargas(cargas_visibles)

    if grado_id and grupo_id and not grupos.filter(id=grupo_id, grado_id=grado_id).exists():
        grupo_id = ''

    grupos_filtrados = grupos
    if grado_id:
        grupos_filtrados = grupos.filter(grado_id=grado_id)

    filtro_asistencias_visibles = Q(asistencias__carga_academica__in=cargas_visibles)

    resumen = filtrar_estudiantes_visibles(Estudiante.objects.filter(
        activo=True
    ), request.user).select_related(
        'grupo',
        'grupo__grado'
    ).annotate(
        total_presentes=Count('asistencias', filter=filtro_asistencias_visibles & Q(asistencias__estado='P')),
        total_ausentes=Count('asistencias', filter=filtro_asistencias_visibles & Q(asistencias__estado='A')),
        total_tardes=Count('asistencias', filter=filtro_asistencias_visibles & Q(asistencias__estado='T')),
        total_justificadas=Count('asistencias', filter=filtro_asistencias_visibles & Q(asistencias__estado='J')),
        total_registros=Count('asistencias', filter=filtro_asistencias_visibles)
    ).order_by('apellidos', 'nombres')

    if grado_id:
        resumen = resumen.filter(grupo__grado_id=grado_id)

    if grupo_id:
        resumen = resumen.filter(grupo_id=grupo_id)

    if busqueda:
        resumen = resumen.filter(
            Q(codigo__icontains=busqueda) |
            Q(documento__icontains=busqueda) |
            Q(nombres__icontains=busqueda) |
            Q(apellidos__icontains=busqueda)
        )

    configuraciones = ConfiguracionAlerta.objects.filter(
        tipo_alerta__nombre='Inasistencia acumulada',
        activa=True
    ).order_by('umbral')

    resumen = list(resumen)
    for estudiante in resumen:
        if estudiante.total_registros > 0:
            estudiante.porcentaje_asistencia = round(
                (estudiante.total_presentes / estudiante.total_registros) * 100, 1
            )
        else:
            estudiante.porcentaje_asistencia = 0

        estudiante.nivel_riesgo = 'NORMAL'

        for config in configuraciones:
            if comparar_valor(estudiante.total_ausentes, config.operador, config.umbral):
                estudiante.nivel_riesgo = config.nivel

    totales = {
        'presentes': sum(estudiante.total_presentes for estudiante in resumen),
        'ausentes': sum(estudiante.total_ausentes for estudiante in resumen),
        'tardes': sum(estudiante.total_tardes for estudiante in resumen),
        'justificadas': sum(estudiante.total_justificadas for estudiante in resumen),
        'registros': sum(estudiante.total_registros for estudiante in resumen),
    }
    totales['porcentaje_asistencia'] = round(
        (totales['presentes'] / totales['registros']) * 100,
        1
    ) if totales['registros'] else 0

    query_params = request.GET.copy()
    query_params.pop('page', None)

    return {
        'resumen': resumen,
        'grados': grados,
        'grupos': grupos_filtrados,
        'todos_grupos': grupos,
        'busqueda': busqueda,
        'grado_id': grado_id,
        'grupo_id': grupo_id,
        'totales': totales,
        'querystring': query_params.urlencode(),
    }


@role_required(puede_gestionar_docencia)
def registrar_asistencia(request):
    estudiantes = []
    carga = None
    fecha = localdate()
    registros_filtrados = []
    asistencia_existente = False

    cargas_disponibles = cargas_visibles_para(request.user)
    form = AsistenciaFiltroForm(cargas=cargas_disponibles)
    filtro_carga = request.GET.get('carga', '')
    filtro_fecha = request.GET.get('fecha', '')

    if request.method == 'POST':
        if 'buscar' in request.POST:
            form = AsistenciaFiltroForm(request.POST, cargas=cargas_disponibles)
            if form.is_valid():
                carga = form.cleaned_data['carga_academica']

                estudiantes = Estudiante.objects.filter(
                    grupo=carga.grupo,
                    activo=True
                ).order_by('apellidos', 'nombres')

                registros_existentes = Asistencia.objects.filter(
                    carga_academica=carga,
                    fecha=fecha
                )

                asistencia_existente = registros_existentes.exists()

                estados_guardados = {
                    registro.estudiante_id: registro
                    for registro in registros_existentes
                }

                for estudiante in estudiantes:
                    if estudiante.id in estados_guardados:
                        estudiante.estado_guardado = estados_guardados[estudiante.id].estado
                        estudiante.observacion_guardada = estados_guardados[estudiante.id].observacion
                    else:
                        estudiante.estado_guardado = 'P'
                        estudiante.observacion_guardada = ''

        elif 'guardar' in request.POST or 'modificar' in request.POST:
            carga_id = request.POST.get('carga_academica')

            try:
                carga = get_object_or_404(cargas_disponibles, id=carga_id)

                estudiantes = Estudiante.objects.filter(
                    grupo=carga.grupo,
                    activo=True
                ).order_by('apellidos', 'nombres')

                registros_existentes = Asistencia.objects.filter(
                    carga_academica=carga,
                    fecha=fecha
                )
                asistencia_existente = registros_existentes.exists()
                registros_previos = {
                    registro.estudiante_id: (registro.estado, registro.observacion or '')
                    for registro in registros_existentes
                }
                cambios_realizados = []
                detalle_cambios = []

                for estudiante in estudiantes:
                    estado = request.POST.get(f'estado_{estudiante.id}', 'P')
                    observacion = request.POST.get(f'observacion_{estudiante.id}', '')
                    previo = registros_previos.get(estudiante.id)
                    accion = 'CREACION' if previo is None else 'EDICION'

                    if previo and (previo[0] != estado or previo[1] != observacion):
                        cambios_realizados.append(
                            f'{estudiante.apellidos} {estudiante.nombres}: {previo[0]} -> {estado}'
                        )
                        detalle_cambios.append({
                            'estudiante': f'{estudiante.apellidos} {estudiante.nombres}',
                            'valor_anterior': previo[0],
                            'valor_nuevo': estado,
                            'observacion_anterior': previo[1],
                            'observacion_nueva': observacion,
                        })

                    Asistencia.objects.update_or_create(
                        estudiante=estudiante,
                        carga_academica=carga,
                        fecha=fecha,
                        defaults={
                            'estado': estado,
                            'observacion': observacion
                        }
                    )

                    if previo is None or previo[0] != estado or previo[1] != observacion:
                        registrar_auditoria_cambio(
                            actor=request.user,
                            tipo='ASISTENCIA',
                            accion=accion,
                            modulo='asistencia',
                            titulo=f'Actualizacion de asistencia en {carga.asignatura.nombre}',
                            descripcion=(
                                f'{request.user.get_full_name() or request.user.username} '
                                f'registro asistencia para {estudiante.apellidos} {estudiante.nombres} '
                                f'el {fecha} en {carga.asignatura.nombre} ({carga.grupo}).'
                            ),
                            estudiante=estudiante,
                            grupo=str(carga.grupo),
                            asignatura=carga.asignatura.nombre,
                            fecha_referencia=fecha,
                            valor_anterior=previo[0] if previo else 'Sin registro',
                            valor_nuevo=estado,
                            referencia_url=f'/asistencia/registrar/?carga={carga.id}',
                            datos_extra={
                                'observacion_anterior': previo[1] if previo else '',
                                'observacion_nueva': observacion,
                            },
                        )

                    evaluar_alertas_academicas(estudiante)

                if 'guardar' in request.POST and not asistencia_existente:
                    messages.success(request, 'Asistencia registrada correctamente para hoy.')
                else:
                    messages.success(request, 'Asistencia modificada correctamente para hoy.')
                    if cambios_realizados:
                        resumen = '; '.join(cambios_realizados[:4])
                        if len(cambios_realizados) > 4:
                            resumen += f'; y {len(cambios_realizados) - 4} cambio(s) mas'
                        crear_notificaciones_docentes(
                            request.user,
                            'ASISTENCIA',
                            f'Modificacion de asistencia en {carga.asignatura.nombre} - {carga.grupo}',
                            (
                                f'Se modificaron {len(cambios_realizados)} registros del {fecha} '
                                f'por {request.user.get_full_name() or request.user.username}. {resumen}.'
                            ),
                            url=f'/asistencia/registrar/?carga={carga.id}',
                            detalle_cambios=detalle_cambios,
                            metadata={
                                'asignatura': carga.asignatura.nombre,
                                'grupo': str(carga.grupo),
                                'fecha': str(fecha),
                            },
                        )

                return redirect(f'/asistencia/registrar/?carga={carga.id}')

            except CargaAcademica.DoesNotExist:
                messages.error(request, 'La carga académica seleccionada no existe.')

    else:
        carga_id = request.GET.get('carga')
        if carga_id:
            try:
                carga = get_object_or_404(cargas_disponibles, id=carga_id)
                form = AsistenciaFiltroForm(initial={
                    'carga_academica': carga
                }, cargas=cargas_disponibles)

                estudiantes = Estudiante.objects.filter(
                    grupo=carga.grupo,
                    activo=True
                ).order_by('apellidos', 'nombres')

                registros_existentes = Asistencia.objects.filter(
                    carga_academica=carga,
                    fecha=fecha
                )

                asistencia_existente = registros_existentes.exists()

                estados_guardados = {
                    registro.estudiante_id: registro
                    for registro in registros_existentes
                }

                for estudiante in estudiantes:
                    if estudiante.id in estados_guardados:
                        estudiante.estado_guardado = estados_guardados[estudiante.id].estado
                        estudiante.observacion_guardada = estados_guardados[estudiante.id].observacion
                    else:
                        estudiante.estado_guardado = 'P'
                        estudiante.observacion_guardada = ''

            except CargaAcademica.DoesNotExist:
                pass

    if filtro_carga and filtro_fecha:
        try:
            registros_filtrados = Asistencia.objects.select_related(
                'estudiante',
                'carga_academica',
                'carga_academica__asignatura',
                'carga_academica__grupo'
            ).filter(
                carga_academica_id=filtro_carga,
                carga_academica__in=cargas_disponibles,
                fecha=filtro_fecha
            ).order_by('estudiante__apellidos', 'estudiante__nombres')
        except Exception:
            registros_filtrados = []

    cargas = cargas_disponibles.select_related('asignatura', 'grupo', 'grupo__grado').order_by(
        'grupo__grado__nombre',
        'grupo__nombre',
        'asignatura__nombre'
    )
    grados, grupos = _grados_y_grupos_desde_cargas(cargas_disponibles)

    return render(request, 'asistencia/registrar_asistencia.html', {
        'form': form,
        'estudiantes': estudiantes,
        'carga': carga,
        'fecha': fecha,
        'asistencia_existente': asistencia_existente,
        'registros_filtrados': registros_filtrados,
        'cargas': cargas,
        'grados': grados,
        'grupos': grupos,
        'filtro_carga': filtro_carga,
        'filtro_fecha': filtro_fecha,
    })


@login_required
@role_required(no_es_estudiante)
def resumen_asistencia(request):
    return render(request, 'asistencia/resumen_asistencia.html', _datos_resumen_asistencia(request))


@login_required
@role_required(no_es_estudiante)
def exportar_resumen_asistencia_excel(request):
    datos = _datos_resumen_asistencia(request)

    libro = Workbook()
    hoja = libro.active
    _preparar_hoja_excel(hoja, 'Resumen asistencia', [
        'Codigo',
        'Documento',
        'Estudiante',
        'Grado',
        'Grupo',
        'Presentes',
        'Ausentes',
        'Tardes',
        'Justificadas',
        'Total registros',
        'Porcentaje asistencia',
        'Nivel de alerta',
    ])

    for estudiante in datos['resumen']:
        hoja.append([
            estudiante.codigo,
            estudiante.documento,
            f'{estudiante.apellidos} {estudiante.nombres}',
            estudiante.grupo.grado.nombre,
            estudiante.grupo.nombre,
            estudiante.total_presentes,
            estudiante.total_ausentes,
            estudiante.total_tardes,
            estudiante.total_justificadas,
            estudiante.total_registros,
            estudiante.porcentaje_asistencia,
            estudiante.nivel_riesgo,
        ])

    _ajustar_columnas(hoja)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="resumen_asistencia.xlsx"'
    libro.save(response)
    return response
