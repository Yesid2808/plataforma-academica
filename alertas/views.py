from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from academico.models import Grado, ReporteAcudiente
from usuarios.permissions import filtrar_alertas_visibles, grupos_visibles_para
from .forms import SeguimientoAlertaForm
from .models import AlertaTemprana, TipoAlerta
from .utils import construir_descripcion_actual_alerta, construir_detalle_alerta


def _alertas_filtradas(request):
    alertas = filtrar_alertas_visibles(
        AlertaTemprana.objects.select_related(
            'estudiante',
            'estudiante__grupo',
            'estudiante__grupo__grado',
            'tipo_alerta',
            'configuracion',
        ).prefetch_related('seguimientos'),
        request.user
    )

    busqueda = request.GET.get('busqueda', '').strip()
    tipo_id = request.GET.get('tipo', '').strip()
    grado_id = request.GET.get('grado', '').strip()
    estado = request.GET.get('estado_alerta', request.GET.get('estado', '')).strip()
    nivel = request.GET.get('nivel_alerta', request.GET.get('nivel', '')).strip()

    if busqueda:
        alertas = alertas.filter(
            Q(estudiante__nombres__icontains=busqueda) |
            Q(estudiante__apellidos__icontains=busqueda) |
            Q(estudiante__codigo__icontains=busqueda) |
            Q(estudiante__documento__icontains=busqueda)
        )

    if tipo_id:
        alertas = alertas.filter(tipo_alerta_id=tipo_id)

    if grado_id:
        alertas = alertas.filter(estudiante__grupo__grado_id=grado_id)

    if estado:
        alertas = alertas.filter(estado=estado)

    if nivel:
        alertas = alertas.filter(nivel=nivel)

    return {
        'queryset': alertas.order_by('-fecha_generacion'),
        'filtro_aplicado': any([busqueda, tipo_id, grado_id, estado, nivel]),
        'busqueda': busqueda,
        'tipo_id': tipo_id,
        'grado_id': grado_id,
        'estado': estado,
        'nivel': nivel,
    }


def _preparar_hoja_excel(hoja, titulo, encabezados):
    hoja.title = titulo
    hoja.append(encabezados)

    fill = PatternFill('solid', fgColor='1F2937')
    for celda in hoja[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = fill


def _ajustar_columnas(hoja):
    for columna in hoja.columns:
        longitud = max(len(str(celda.value or '')) for celda in columna)
        hoja.column_dimensions[get_column_letter(columna[0].column)].width = min(longitud + 2, 45)


def _nivel_legible(nivel):
    return {
        'ATENCION': 'Atencion',
        'RIESGO': 'Riesgo',
        'CRITICO': 'Critico',
    }.get(nivel, nivel)


@login_required
def lista_alertas(request):
    from academico.views import gestion_seguimiento
    return gestion_seguimiento(request)


@login_required
def detalle_alerta(request, pk):
    alerta = get_object_or_404(
        filtrar_alertas_visibles(AlertaTemprana.objects.select_related(
            'estudiante',
            'estudiante__grupo',
            'estudiante__grupo__grado',
            'tipo_alerta',
            'configuracion'
        ), request.user),
        pk=pk
    )
    seguimientos = alerta.seguimientos.select_related('registrado_por')

    if request.method == 'POST':
        form = SeguimientoAlertaForm(request.POST)
        if form.is_valid():
            seguimiento = form.save(commit=False)
            seguimiento.alerta = alerta
            seguimiento.registrado_por = request.user
            seguimiento.save()

            if seguimiento.resultado == 'CERRADO':
                alerta.estado = 'CERRADA'
            elif alerta.estado == 'ACTIVA':
                alerta.estado = 'REVISADA'
            alerta.save(update_fields=['estado'])

            messages.success(request, 'Seguimiento registrado correctamente.')
            return redirect('detalle_alerta', pk=alerta.pk)
    else:
        form = SeguimientoAlertaForm()

    return render(request, 'alertas/detalle_alerta.html', {
        'alerta': alerta,
        'seguimientos': seguimientos,
        'form': form,
        'detalle_alerta': construir_detalle_alerta(alerta),
    })


@login_required
def exportar_alertas_excel(request):
    alertas = _alertas_filtradas(request)['queryset']

    libro = Workbook()
    hoja = libro.active
    _preparar_hoja_excel(hoja, 'Alertas', [
        'Codigo',
        'Documento',
        'Estudiante',
        'Grado',
        'Grupo',
        'Tipo de alerta',
        'Nivel',
        'Estado',
        'Fecha de generacion',
        'Descripcion',
        'Ultima accion',
        'Resultado seguimiento',
        'Proxima revision',
    ])

    for alerta in alertas:
        estudiante = alerta.estudiante
        ultimo_seguimiento = next(iter(alerta.seguimientos.all()), None)
        hoja.append([
            estudiante.codigo,
            estudiante.documento,
            f'{estudiante.apellidos} {estudiante.nombres}',
            estudiante.grupo.grado.nombre,
            estudiante.grupo.nombre,
            alerta.tipo_alerta.nombre if alerta.tipo_alerta else '',
            _nivel_legible(alerta.nivel),
            alerta.get_estado_display(),
            alerta.fecha_generacion.strftime('%Y-%m-%d %H:%M'),
            construir_descripcion_actual_alerta(alerta),
            ultimo_seguimiento.get_accion_display() if ultimo_seguimiento else '',
            ultimo_seguimiento.get_resultado_display() if ultimo_seguimiento else '',
            ultimo_seguimiento.proxima_revision.strftime('%Y-%m-%d') if ultimo_seguimiento and ultimo_seguimiento.proxima_revision else '',
        ])

    _ajustar_columnas(hoja)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="alertas_tempranas.xlsx"'
    libro.save(response)
    return response


@require_POST
@login_required
def cambiar_estado_alerta(request, pk, nuevo_estado):
    alerta = get_object_or_404(
        filtrar_alertas_visibles(AlertaTemprana.objects.all(), request.user),
        pk=pk
    )

    estados_validos = ['ACTIVA', 'REVISADA', 'CERRADA']
    if nuevo_estado not in estados_validos:
        messages.error(request, 'Estado no valido.')
        return redirect('lista_alertas')

    alerta.estado = nuevo_estado
    alerta.save()

    if nuevo_estado == 'REVISADA':
        messages.success(request, 'La alerta fue marcada como revisada.')
    elif nuevo_estado == 'CERRADA':
        messages.success(request, 'La alerta fue cerrada correctamente.')
    else:
        messages.success(request, 'La alerta fue actualizada.')

    return redirect('lista_alertas')
