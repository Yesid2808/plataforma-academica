from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.db.models import Avg
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import localdate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from academico.models import Estudiante, HorarioClase, PeriodoAcademico
from asistencia.models import Asistencia
from alertas.models import AlertaTemprana
from evaluacion.models import Calificacion
from usuarios.models import AuditoriaCambio
from usuarios.permissions import (
    cargas_visibles_para,
    es_estudiante,
    filtrar_alertas_visibles,
    filtrar_estudiantes_visibles,
    obtener_estudiante_usuario,
)


def _preparar_hoja_excel(hoja, titulo, encabezados):
    hoja.title = titulo
    hoja.append(encabezados)

    fill = PatternFill('solid', fgColor='1F2937')
    for celda in hoja[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = fill


def _ajustar_columnas(hoja):
    for columna in hoja.columns:
        longitud = max(len(str(celda.value or '')) for celda in columna)
        hoja.column_dimensions[get_column_letter(columna[0].column)].width = min(longitud + 2, 45)


def _construir_powerbi_embed_url(url):
    if not url:
        return ''

    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query.setdefault('pageView', 'fitToWidth')
    query.setdefault('navContentPaneEnabled', 'false')
    query.setdefault('filterPaneEnabled', 'false')
    query.setdefault('chromeless', '1')
    return urlunparse(parsed._replace(query=urlencode(query)))


@login_required
def inicio(request):
    es_docente = getattr(request.user, 'rol', '') == 'DOC'
    es_estudiante_dashboard = es_estudiante(request.user)
    estudiante_actual = obtener_estudiante_usuario(request.user)
    hoy = localdate()
    periodo_activo = PeriodoAcademico.objects.select_related('anio_lectivo').filter(
        activo=True,
        fecha_inicio__lte=hoy,
        fecha_fin__gte=hoy,
    ).order_by('numero').first()
    cargas_visibles = cargas_visibles_para(request.user)
    estudiantes_visibles = filtrar_estudiantes_visibles(Estudiante.objects.filter(activo=True), request.user)
    alertas_visibles = filtrar_alertas_visibles(AlertaTemprana.objects.all(), request.user)
    asistencias_visibles = Asistencia.objects.filter(carga_academica__in=cargas_visibles)
    calificaciones_visibles = Calificacion.objects.filter(actividad__carga_academica__in=cargas_visibles)

    if es_estudiante_dashboard:
        if estudiante_actual:
            asistencias_visibles = asistencias_visibles.filter(estudiante=estudiante_actual)
            calificaciones_visibles = calificaciones_visibles.filter(estudiante=estudiante_actual)
        else:
            asistencias_visibles = asistencias_visibles.none()
            calificaciones_visibles = calificaciones_visibles.none()

    total_estudiantes = estudiantes_visibles.count()
    total_grupos = cargas_visibles.values('grupo_id').distinct().count()
    total_asignaturas = cargas_visibles.values('asignatura_id').distinct().count()

    total_asistencias = asistencias_visibles.count()
    total_presentes = asistencias_visibles.filter(estado='P').count()
    total_ausentes = asistencias_visibles.filter(estado='A').count()
    total_tardes = asistencias_visibles.filter(estado='T').count()
    total_justificadas = asistencias_visibles.filter(estado='J').count()

    if total_asistencias > 0:
        porcentaje_asistencia = round((total_presentes / total_asistencias) * 100, 1)
    else:
        porcentaje_asistencia = 0

    total_alertas_activas = alertas_visibles.filter(estado='ACTIVA').count()
    total_alertas_criticas = alertas_visibles.filter(
        estado='ACTIVA',
        nivel='CRITICO'
    ).count()
    total_docentes = cargas_visibles.values('docente_id').distinct().count()

    total_calificaciones = calificaciones_visibles.count()
    total_bajo_desempeno = calificaciones_visibles.filter(nota__lt=3).count()
    porcentaje_bajo_desempeno = round(
        (total_bajo_desempeno / total_calificaciones) * 100,
        1
    ) if total_calificaciones else 0
    promedio_general = round(
        calificaciones_visibles.aggregate(promedio=Avg('nota'))['promedio'] or 0,
        2
    )
    calificaciones_altas = calificaciones_visibles.filter(nota__gte=4).count()
    calificaciones_medias = calificaciones_visibles.filter(nota__gte=3, nota__lt=4).count()

    alertas_destacadas = alertas_visibles.select_related(
        'estudiante',
        'estudiante__grupo',
        'tipo_alerta'
    ).filter(
        estado='ACTIVA'
    ).order_by('-fecha_generacion')[:5]

    resumen_estudiantes = list(
        estudiantes_visibles.select_related(
            'grupo',
            'grupo__grado'
        ).annotate(
            total_ausentes=Count('asistencias', filter=Q(asistencias__estado='A')),
            total_presentes=Count('asistencias', filter=Q(asistencias__estado='P')),
            total_tardes=Count('asistencias', filter=Q(asistencias__estado='T')),
            total_justificadas=Count('asistencias', filter=Q(asistencias__estado='J')),
            total_registros=Count('asistencias')
        ).order_by('-total_ausentes', 'apellidos', 'nombres')
    )

    for estudiante in resumen_estudiantes:
        if estudiante.total_registros > 0:
            estudiante.porcentaje_asistencia = round(
                (estudiante.total_presentes / estudiante.total_registros) * 100, 1
            )
        else:
            estudiante.porcentaje_asistencia = 0

        if estudiante.total_ausentes >= 3:
            estudiante.nivel_riesgo = 'RIESGO'
        elif estudiante.total_ausentes == 2:
            estudiante.nivel_riesgo = 'ATENCION'
        else:
            estudiante.nivel_riesgo = 'NORMAL'

    top_ausencias = [e for e in resumen_estudiantes if e.total_ausentes > 0][:5]
    bajo_rendimiento_estudiantes = list(
        estudiantes_visibles.annotate(
            promedio_general=Avg('calificaciones__nota')
        ).filter(
            promedio_general__isnull=False,
            promedio_general__lt=3
        ).select_related('grupo', 'grupo__grado').order_by('promedio_general', 'apellidos', 'nombres')[:5]
    )

    resumen_grupos = {}
    for estudiante in resumen_estudiantes:
        clave = str(estudiante.grupo)
        info = resumen_grupos.setdefault(clave, {
            'grupo': clave,
            'estudiantes': 0,
            'ausentes': 0,
            'presentes': 0,
            'registros': 0,
        })
        info['estudiantes'] += 1
        info['ausentes'] += estudiante.total_ausentes
        info['presentes'] += estudiante.total_presentes
        info['registros'] += estudiante.total_registros

    resumen_grupos = list(resumen_grupos.values())
    for grupo in resumen_grupos:
        grupo['porcentaje_asistencia'] = round(
            (grupo['presentes'] / grupo['registros']) * 100,
            1
        ) if grupo['registros'] else 0

    grupos_alerta = sorted(
        resumen_grupos,
        key=lambda item: (-item['ausentes'], item['grupo'])
    )[:6]

    tendencia_labels = []
    tendencia_asistencia = []
    for desplazamiento in range(6, -1, -1):
        fecha_revision = hoy.fromordinal(hoy.toordinal() - desplazamiento)
        if fecha_revision.weekday() >= 5:
            continue
        registros_dia = asistencias_visibles.filter(fecha=fecha_revision)
        total_dia = registros_dia.count()
        presentes_dia = registros_dia.filter(estado='P').count()
        porcentaje_dia = round((presentes_dia / total_dia) * 100, 1) if total_dia else 0
        tendencia_labels.append(fecha_revision.strftime('%d/%m'))
        tendencia_asistencia.append(porcentaje_dia)

    dia_consulta = hoy.isoweekday()
    if dia_consulta > 5:
        dia_consulta = 1

    bloques_hoy = list(HorarioClase.objects.select_related(
        'carga_academica',
        'carga_academica__grupo',
        'carga_academica__grupo__grado',
        'carga_academica__asignatura',
        'carga_academica__docente',
    ).filter(
        dia_semana=dia_consulta,
        carga_academica__in=cargas_visibles,
    ).order_by('hora_inicio', 'carga_academica__grupo__grado__nombre', 'carga_academica__grupo__nombre')[:8])

    proximas_acciones = []
    if es_docente:
        proximas_acciones = [
            {'label': 'Jornada docente', 'url': '/asistencia/jornada/', 'icon': 'bi-easel2-fill'},
            {'label': 'Registrar asistencia', 'url': '/asistencia/registrar/', 'icon': 'bi-calendar-check-fill'},
            {'label': 'Calificaciones', 'url': '/evaluacion/actividades/', 'icon': 'bi-journal-check'},
            {'label': 'Mis notificaciones', 'url': '/usuarios/notificaciones/', 'icon': 'bi-bell-fill'},
        ]
    else:
        proximas_acciones = [
            {'label': 'Seguimiento academico', 'url': '/academico/seguimiento/', 'icon': 'bi-activity'},
            {'label': 'Notificaciones', 'url': '/usuarios/notificaciones/', 'icon': 'bi-bell-fill'},
            {'label': 'Catalogos', 'url': '/academico/gestion/', 'icon': 'bi-building-fill-gear'},
        ]
    if es_estudiante_dashboard:
        proximas_acciones = [
            {'label': 'Mi horario', 'url': '/academico/horarios/', 'icon': 'bi-calendar3-week-fill'},
            {'label': 'Mi asistencia', 'url': '/asistencia/resumen/', 'icon': 'bi-calendar-check-fill'},
            {'label': 'Mis notas', 'url': '/evaluacion/resumen/', 'icon': 'bi-journal-check'},
            {'label': 'Mi seguimiento', 'url': '/academico/mi-seguimiento/', 'icon': 'bi-activity'},
        ]

    actividad_reciente = AuditoriaCambio.objects.select_related('actor')[:6]
    salud_operativa = 100
    if total_alertas_activas:
        salud_operativa -= min(35, total_alertas_activas * 2)
    if porcentaje_asistencia < 95:
        salud_operativa -= min(25, int(95 - porcentaje_asistencia))
    if porcentaje_bajo_desempeno > 0:
        salud_operativa -= min(25, int(porcentaje_bajo_desempeno / 2))
    salud_operativa = max(35, min(100, salud_operativa))

    context = {
        'fecha_panel': hoy,
        'periodo_activo': periodo_activo,
        'es_docente_dashboard': es_docente,
        'es_estudiante_dashboard': es_estudiante_dashboard,
        'estudiante_actual': estudiante_actual,
        'total_estudiantes': total_estudiantes,
        'total_grupos': total_grupos,
        'total_asignaturas': total_asignaturas,
        'total_docentes': total_docentes,
        'total_asistencias': total_asistencias,
        'total_presentes': total_presentes,
        'total_ausentes': total_ausentes,
        'total_tardes': total_tardes,
        'total_justificadas': total_justificadas,
        'porcentaje_asistencia': porcentaje_asistencia,
        'total_alertas_activas': total_alertas_activas,
        'total_alertas_criticas': total_alertas_criticas,
        'total_calificaciones': total_calificaciones,
        'total_bajo_desempeno': total_bajo_desempeno,
        'porcentaje_bajo_desempeno': porcentaje_bajo_desempeno,
        'promedio_general': promedio_general,
        'calificaciones_altas': calificaciones_altas,
        'calificaciones_medias': calificaciones_medias,
        'salud_operativa': salud_operativa,
        'alertas_destacadas': alertas_destacadas,
        'top_ausencias': top_ausencias,
        'bajo_rendimiento_estudiantes': bajo_rendimiento_estudiantes,
        'bloques_hoy': bloques_hoy,
        'total_bloques_hoy': len(bloques_hoy),
        'dia_horario': dict(HorarioClase.DIA_CHOICES).get(dia_consulta, 'Lunes'),
        'powerbi_dashboard_url': settings.POWERBI_DASHBOARD_URL,
        'powerbi_dashboard_embed_url': _construir_powerbi_embed_url(settings.POWERBI_DASHBOARD_URL),
        'proximas_acciones': proximas_acciones,
        'actividad_reciente': actividad_reciente,
        'asistencia_chart_labels': ['Presentes', 'Ausentes', 'Tardes', 'Justificadas'],
        'asistencia_chart_data': [total_presentes, total_ausentes, total_tardes, total_justificadas],
        'rendimiento_chart_labels': ['En riesgo (< 3.0)', 'Estable (3.0 - 3.99)', 'Alto (4.0 - 5.0)'],
        'rendimiento_chart_data': [total_bajo_desempeno, calificaciones_medias, calificaciones_altas],
        'grupos_chart_labels': [item['grupo'] for item in grupos_alerta],
        'grupos_chart_data': [item['ausentes'] for item in grupos_alerta],
        'tendencia_chart_labels': tendencia_labels,
        'tendencia_chart_data': tendencia_asistencia,
        'grupos_alerta': grupos_alerta,
    }
    return render(request, 'dashboard/inicio.html', context)


@login_required
def exportar_estudiantes_riesgo_excel(request):
    alertas = filtrar_alertas_visibles(
        AlertaTemprana.objects.select_related(
            'estudiante',
            'estudiante__grupo',
            'estudiante__grupo__grado',
            'tipo_alerta',
        ).filter(estado='ACTIVA'),
        request.user
    ).order_by('estudiante__apellidos', 'estudiante__nombres', '-nivel')

    filas = {}
    for alerta in alertas:
        estudiante = alerta.estudiante
        fila = filas.setdefault(estudiante.id, {
            'estudiante': estudiante,
            'total': 0,
            'criticas': 0,
            'riesgo': 0,
            'atencion': 0,
            'tipos': set(),
        })
        fila['total'] += 1
        fila['tipos'].add(alerta.tipo_alerta.nombre if alerta.tipo_alerta else 'Sin tipo')

        if alerta.nivel == 'CRITICO':
            fila['criticas'] += 1
        elif alerta.nivel == 'RIESGO':
            fila['riesgo'] += 1
        elif alerta.nivel == 'ATENCION':
            fila['atencion'] += 1

    libro = Workbook()
    hoja = libro.active
    _preparar_hoja_excel(hoja, 'Estudiantes en riesgo', [
        'Codigo',
        'Documento',
        'Estudiante',
        'Grado',
        'Grupo',
        'Alertas activas',
        'Criticas',
        'Riesgo',
        'Atencion',
        'Tipos de alerta',
    ])

    for fila in filas.values():
        estudiante = fila['estudiante']
        hoja.append([
            estudiante.codigo,
            estudiante.documento,
            f'{estudiante.apellidos} {estudiante.nombres}',
            estudiante.grupo.grado.nombre,
            estudiante.grupo.nombre,
            fila['total'],
            fila['criticas'],
            fila['riesgo'],
            fila['atencion'],
            ', '.join(sorted(fila['tipos'])),
        ])

    _ajustar_columnas(hoja)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="estudiantes_en_riesgo.xlsx"'
    libro.save(response)
    return response
