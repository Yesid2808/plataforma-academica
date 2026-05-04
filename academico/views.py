from io import BytesIO
from datetime import datetime, timedelta
import re

from django.core.paginator import Paginator
from django.db.models import Count, Exists, OuterRef, Q

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from usuarios.decorators import role_required
from usuarios.permissions import (
    cargas_visibles_para,
    es_docente,
    es_estudiante,
    filtrar_alertas_visibles,
    filtrar_estudiantes_visibles,
    grupos_visibles_para,
    obtener_estudiante_usuario,
    puede_gestionar_docencia,
    puede_gestionar_catalogos,
    puede_ver_todo,
)
from usuarios.models import Usuario
from .reportes import (
    construir_datos_reporte_estudiante,
    enviar_reportes_estudiantes,
    enviar_reporte_estudiante_por_correo,
    generar_excel_reporte_estudiante,
    obtener_estado_correo,
    PERIODOS_REPORTE,
    probar_conexion_correo,
    registrar_reporte,
)
from .utils import inferir_genero_por_nombre
from .forms import (
    AnioLectivoForm,
    AsignaturaForm,
    CargaAcademicaForm,
    EstudianteForm,
    EstudianteUpdateForm,
    GradoForm,
    GrupoForm,
    ImportarEstudiantesForm,
    PeriodoAcademicoForm,
)
from .models import (
    AnioLectivo,
    Asignatura,
    CargaAcademica,
    Estudiante,
    Grado,
    Grupo,
    HorarioClase,
    PeriodoAcademico,
    ReporteAcudiente,
)
from alertas.models import AlertaTemprana
from alertas.utils import construir_descripcion_actual_alerta


CATALOGOS = {
    'anios': {
        'titulo': 'Años lectivos',
        'singular': 'año lectivo',
        'model': AnioLectivo,
        'form': AnioLectivoForm,
        'queryset': lambda: AnioLectivo.objects.all().order_by('-anio'),
        'columns': [('Año', 'anio'), ('Activo', 'activo')],
    },
    'periodos': {
        'titulo': 'Periodos académicos',
        'singular': 'periodo académico',
        'model': PeriodoAcademico,
        'form': PeriodoAcademicoForm,
        'queryset': lambda: PeriodoAcademico.objects.select_related('anio_lectivo').all().order_by('anio_lectivo', 'numero'),
        'columns': [('Nombre', 'nombre'), ('Número', 'numero'), ('Año lectivo', 'anio_lectivo'), ('Activo', 'activo')],
    },
    'grados': {
        'titulo': 'Grados',
        'singular': 'grado',
        'model': Grado,
        'form': GradoForm,
        'queryset': lambda: Grado.objects.all().order_by('nombre'),
        'columns': [('Nombre', 'nombre')],
    },
    'grupos': {
        'titulo': 'Grupos',
        'singular': 'grupo',
        'model': Grupo,
        'form': GrupoForm,
        'queryset': lambda: Grupo.objects.select_related('grado', 'director_grupo').all().order_by('grado__nombre', 'nombre'),
        'columns': [('Nombre', 'nombre'), ('Grado', 'grado'), ('Director', 'director_grupo'), ('Activo', 'activo')],
    },
    'asignaturas': {
        'titulo': 'Asignaturas',
        'singular': 'asignatura',
        'model': Asignatura,
        'form': AsignaturaForm,
        'queryset': lambda: Asignatura.objects.all().order_by('nombre'),
        'columns': [('Nombre', 'nombre'), ('Intensidad horaria', 'intensidad_horaria'), ('Activa', 'activa')],
    },
    'cargas': {
        'titulo': 'Cargas académicas',
        'singular': 'carga académica',
        'model': CargaAcademica,
        'form': CargaAcademicaForm,
        'queryset': lambda: CargaAcademica.objects.select_related('docente', 'grupo', 'asignatura', 'anio_lectivo').all().order_by('grupo__grado__nombre', 'grupo__nombre', 'asignatura__nombre'),
        'columns': [('Docente', 'docente'), ('Grupo', 'grupo'), ('Asignatura', 'asignatura'), ('Año lectivo', 'anio_lectivo'), ('Activa', 'activo')],
    },
}


def _normalizar_numero(valor):
    return re.sub(r'\D', '', str(valor or ''))


def _parsear_fecha_query(valor):
    if not valor:
        return None
    try:
        return datetime.strptime(valor, '%Y-%m-%d').date()
    except ValueError:
        return None


def _page_window(page_obj, max_visible=5):
    total_pages = page_obj.paginator.num_pages
    current_page = page_obj.number

    if total_pages <= max_visible:
        start_page = 1
        end_page = total_pages
    else:
        half_window = max_visible // 2
        start_page = max(current_page - half_window, 1)
        end_page = start_page + max_visible - 1

        if end_page > total_pages:
            end_page = total_pages
            start_page = max(end_page - max_visible + 1, 1)

    return list(range(start_page, end_page + 1))


def _estudiantes_visibles_qs(user):
    return filtrar_estudiantes_visibles(
        Estudiante.objects.select_related('grupo', 'grupo__grado', 'grupo__director_grupo'),
        user,
    )


def _obtener_next_seguro(request, fallback='lista_estudiantes'):
    next_url = request.GET.get('next') or request.POST.get('next') or ''
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return reverse(fallback)


def _estudiante_sesion_o_404(user):
    estudiante = obtener_estudiante_usuario(user)
    if not estudiante:
        raise Http404('No hay un estudiante asociado a esta cuenta.')
    return estudiante


def _tiene_filtros_estudiantes(request):
    return request.GET.get('filtrar') == '1'


def _resolver_grupo_importacion(valor):
    if valor in (None, ''):
        raise Grupo.DoesNotExist

    texto = str(valor).strip()
    if texto.isdigit():
        return Grupo.objects.select_related('grado').get(id=int(texto))

    if '-' in texto:
        grado_nombre, grupo_nombre = texto.split('-', 1)
        return Grupo.objects.select_related('grado').get(
            grado__nombre__iexact=grado_nombre.strip(),
            nombre__iexact=grupo_nombre.strip(),
        )

    grupos = Grupo.objects.select_related('grado').filter(nombre__iexact=texto)
    if grupos.count() == 1:
        return grupos.first()

    raise Grupo.DoesNotExist


def _resolver_grupo_por_grado_y_grupo(grado_nombre, grupo_nombre):
    if not grado_nombre or not grupo_nombre:
        raise Grupo.DoesNotExist

    return Grupo.objects.select_related('grado').get(
        grado__nombre__iexact=str(grado_nombre).strip(),
        nombre__iexact=str(grupo_nombre).strip(),
    )


def _parsear_fecha_excel(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if hasattr(valor, 'year') and hasattr(valor, 'month') and hasattr(valor, 'day'):
        return valor
    if not valor:
        return None

    texto = str(valor).strip()
    for formato in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _obtener_catalogo(catalogo):
    try:
        return CATALOGOS[catalogo]
    except KeyError as exc:
        raise Http404('Catálogo académico no encontrado.') from exc


def _formatear_fila(obj, columns):
    valores = []
    for label, attr in columns:
        value = getattr(obj, attr)
        if isinstance(value, bool):
            value = 'Sí' if value else 'No'
        valores.append((label, value))
    return valores


def _estudiantes_filtrados(request):
    busqueda = request.GET.get('busqueda', '').strip()
    grado_id = request.GET.get('grado', '').strip()
    grupo_id = request.GET.get('grupo', '').strip()
    estado = request.GET.get('estado', '').strip()

    if not _tiene_filtros_estudiantes(request):
        return Estudiante.objects.none(), {
            'busqueda': busqueda,
            'grado_id': grado_id,
            'grupo_id': grupo_id,
            'estado': estado,
            'filtrar': '',
            'filtros_aplicados': False,
        }

    estudiantes = _estudiantes_visibles_qs(request.user)

    if grado_id and grupo_id:
        grupo_valido = Grupo.objects.filter(id=grupo_id, grado_id=grado_id).exists()
        if not grupo_valido:
            grupo_id = ''

    if busqueda:
        estudiantes = estudiantes.filter(
            Q(codigo__icontains=busqueda) |
            Q(documento__icontains=busqueda) |
            Q(nombres__icontains=busqueda) |
            Q(apellidos__icontains=busqueda)
        )

    if grado_id:
        estudiantes = estudiantes.filter(grupo__grado_id=grado_id)

    if grupo_id:
        estudiantes = estudiantes.filter(grupo_id=grupo_id)

    if estado == 'activo':
        estudiantes = estudiantes.filter(activo=True)
    elif estado == 'inactivo':
        estudiantes = estudiantes.filter(activo=False)

    return estudiantes.order_by('apellidos', 'nombres'), {
        'busqueda': busqueda,
        'grado_id': grado_id,
        'grupo_id': grupo_id,
        'estado': estado,
        'filtrar': '1',
        'filtros_aplicados': True,
    }


def _estudiantes_reportes_filtrados(request):
    estudiantes_base = filtrar_estudiantes_visibles(
        Estudiante.objects.select_related('grupo', 'grupo__grado').filter(activo=True),
        request.user
    )

    busqueda = request.GET.get('busqueda', '').strip()
    grado_id = request.GET.get('grado', '').strip()
    grupo_id = request.GET.get('grupo', '').strip()
    periodo = request.GET.get('periodo', 'semanal').strip()
    estado = request.GET.get('estado', '').strip().lower()
    filtrar = request.GET.get('filtrar', '').strip() == '1'

    grupos = Grupo.objects.select_related('grado').filter(
        id__in=estudiantes_base.values_list('grupo_id', flat=True).distinct()
    ).order_by('grado__nombre', 'nombre')
    grados = Grado.objects.filter(grupos__in=grupos).distinct().order_by('nombre')

    if grado_id and grupo_id and not grupos.filter(id=grupo_id, grado_id=grado_id).exists():
        grupo_id = ''

    periodo = periodo if periodo in PERIODOS_REPORTE else 'semanal'

    todos_grupos = Grupo.objects.select_related('grado').filter(
        id__in=filtrar_estudiantes_visibles(
            Estudiante.objects.filter(activo=True), request.user
        ).values_list('grupo_id', flat=True).distinct()
    ).order_by('grado__nombre', 'nombre')

    if not filtrar:
        return estudiantes_base.none(), {
            'busqueda': busqueda,
            'grado_id': grado_id,
            'grupo_id': grupo_id,
            'periodo': periodo,
            'estado': estado,
            'filtrar': filtrar,
            'filtros_aplicados': False,
            'grados': grados,
            'grupos': grupos,
            'todos_grupos': todos_grupos,
        }

    estudiantes = estudiantes_base

    if busqueda:
        estudiantes = estudiantes.filter(
            Q(codigo__icontains=busqueda) |
            Q(documento__icontains=busqueda) |
            Q(nombres__icontains=busqueda) |
            Q(apellidos__icontains=busqueda) |
            Q(acudiente__icontains=busqueda) |
            Q(correo_acudiente__icontains=busqueda)
        )

    if grado_id:
        estudiantes = estudiantes.filter(grupo__grado_id=grado_id)
        grupos = grupos.filter(grado_id=grado_id)

    if grupo_id:
        estudiantes = estudiantes.filter(grupo_id=grupo_id)

    reporte_enviado_subquery = ReporteAcudiente.objects.filter(
        estudiante_id=OuterRef('pk'),
        periodo=periodo,
        estado='ENVIADO',
    )
    estudiantes = estudiantes.annotate(tiene_reporte_enviado=Exists(reporte_enviado_subquery))

    if estado == 'enviado':
        estudiantes = estudiantes.filter(tiene_reporte_enviado=True)
    elif estado == 'pendiente':
        estudiantes = estudiantes.filter(tiene_reporte_enviado=False)

    return estudiantes.order_by('apellidos', 'nombres'), {
        'busqueda': busqueda,
        'grado_id': grado_id,
        'grupo_id': grupo_id,
        'periodo': periodo,
        'estado': estado,
        'filtrar': filtrar,
        'filtros_aplicados': True,
        'grados': grados,
        'grupos': grupos,
        'todos_grupos': todos_grupos,
    }


@login_required
def lista_estudiantes(request):
    estudiantes, filtros = _estudiantes_filtrados(request)

    paginator = Paginator(estudiantes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    grupos = grupos_visibles_para(request.user).select_related('director_grupo').order_by('grado__nombre', 'nombre')
    grados = Grado.objects.filter(
        id__in=grupos.values_list('grado_id', flat=True).distinct()
    ).order_by('nombre')
    grupos_filtrados = grupos
    if filtros['grado_id']:
        grupos_filtrados = grupos.filter(grado_id=filtros['grado_id'])

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    context = {
        'page_obj': page_obj,
        'grados': grados,
        'grupos': grupos_filtrados,
        'todos_grupos': grupos,
        'busqueda': filtros['busqueda'],
        'grado_id': filtros['grado_id'],
        'grupo_id': filtros['grupo_id'],
        'estado': filtros['estado'],
        'filtrar': filtros['filtrar'],
        'filtros_aplicados': filtros['filtros_aplicados'],
        'querystring': querystring,
        'page_numbers': _page_window(page_obj),
        'page_param': 'page',
        'current_path': request.get_full_path(),
        'can_manage_students': puede_ver_todo(request.user),
    }
    return render(request, 'academico/lista_estudiantes.html', context)

@role_required(puede_ver_todo)
def reactivar_estudiante(request, pk):
    estudiante = get_object_or_404(_estudiantes_visibles_qs(request.user), pk=pk)

    if request.method == 'POST':
        estudiante.activo = True
        estudiante.save()
        messages.success(
            request,
            f'Estudiante {estudiante.nombres} {estudiante.apellidos} reactivado correctamente.'
        )
        return redirect('lista_estudiantes')

    return render(request, 'academico/reactivar_estudiante.html', {'estudiante': estudiante})


@role_required(puede_ver_todo)
def descargar_plantilla_estudiantes(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla Estudiantes'

    encabezados = [
        'tipo_documento',
        'documento',
        'nombres',
        'apellidos',
        'genero',
        'fecha_nacimiento',
        'grado',
        'grupo',
        'grupo_id',
        'correo',
        'whatsapp',
        'acudiente',
        'correo_acudiente',
        'telefono_acudiente',
        'whatsapp_acudiente',
        'direccion',
        'activo',
    ]

    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.append([
        'TI',
        '1032456789',
        'Juan',
        'Pérez',
        'M',
        '2010-05-12',
        '8',
        'A',
        '1',
        'juan@email.com',
        '3001234567',
        'Ana Pérez',
        'ana@email.com',
        '3001234567',
        '3001234567',
        'Calle 1',
        'true'
    ])

    ws.append([
        'TI',
        '1045678901',
        'María',
        'Gómez',
        'F',
        '2011-07-03',
        '8',
        'B',
        '1',
        'maria@email.com',
        '3009876543',
        'Luis Gómez',
        'luis@email.com',
        '3009876543',
        '3009876543',
        'Calle 2',
        'true'
    ])

    ws_instrucciones = wb.create_sheet('Instrucciones')
    ws_instrucciones.append(['Campo', 'Obligatorio', 'Descripcion', 'Ejemplo'])
    for cell in ws_instrucciones[1]:
        cell.font = Font(bold=True)

    for fila in [
        ('tipo_documento', 'Si', 'Usa TI, CC, RC o CE.', 'TI'),
        ('documento', 'Si', 'Solo numeros, sin puntos, minimo 8 digitos.', '1032456789'),
        ('nombres', 'Si', 'Nombres completos del estudiante.', 'Juan David'),
        ('apellidos', 'Si', 'Apellidos completos del estudiante.', 'Perez Gomez'),
        ('genero', 'Si', 'Solo M o F.', 'M'),
        ('fecha_nacimiento', 'Si', 'Formato YYYY-MM-DD o fecha valida de Excel.', '2010-05-12'),
        ('grado', 'Si', 'Grado academico del estudiante.', '8'),
        ('grupo', 'Si', 'Grupo academico del estudiante.', 'A'),
        ('grupo_id', 'Si', 'Id del grupo segun la hoja Grupos. Debe coincidir con grado y grupo.', '1'),
        ('correo', 'Si', 'Correo del estudiante.', 'juan@email.com'),
        ('whatsapp', 'Si', 'Solo numeros.', '3001234567'),
        ('acudiente', 'Si', 'Nombre completo del acudiente.', 'Ana Maria Perez'),
        ('correo_acudiente', 'Si', 'Correo del acudiente.', 'ana@email.com'),
        ('telefono_acudiente', 'Si', 'Telefono del acudiente. Solo numeros.', '3001234567'),
        ('whatsapp_acudiente', 'Si', 'WhatsApp del acudiente. Solo numeros.', '3001234567'),
        ('direccion', 'Si', 'Direccion del estudiante.', 'Calle 45 # 12-30'),
        ('activo', 'Si', 'Acepta true/false, si/no, 1/0 o activo/inactivo.', 'true'),
    ]:
        ws_instrucciones.append(list(fila))

    ws_grupos = wb.create_sheet('Grupos')
    ws_grupos.append(['grupo_id', 'grado', 'grupo', 'nombre_mostrado', 'director_grupo'])
    for cell in ws_grupos[1]:
        cell.font = Font(bold=True)

    for grupo in Grupo.objects.select_related('grado', 'director_grupo').filter(activo=True).order_by('grado__nombre', 'nombre'):
        ws_grupos.append([
            grupo.id,
            grupo.grado.nombre,
            grupo.nombre,
            str(grupo),
            grupo.director_grupo.get_full_name() if grupo.director_grupo else '',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_estudiantes.xlsx'
    return response

@role_required(puede_ver_todo)
def crear_estudiante(request):
    if request.method == 'POST':
        form = EstudianteForm(request.POST, request.FILES)
        if form.is_valid():
            estudiante = form.save()
            messages.success(request, f'Estudiante registrado correctamente con código {estudiante.codigo}.')
            return redirect('lista_estudiantes')
    else:
        form = EstudianteForm()

    return render(request, 'academico/crear_estudiante.html', {'form': form})


@role_required(puede_ver_todo)
def editar_estudiante(request, pk):
    estudiante = get_object_or_404(_estudiantes_visibles_qs(request.user), pk=pk)

    if request.method == 'POST':
        form = EstudianteUpdateForm(request.POST, request.FILES, instance=estudiante)
        if form.is_valid():
            form.save()
            messages.success(request, 'Estudiante actualizado correctamente.')
            return redirect('lista_estudiantes')
    else:
        form = EstudianteUpdateForm(instance=estudiante)

    return render(request, 'academico/editar_estudiante.html', {
        'form': form,
        'estudiante': estudiante,
    })


@role_required(puede_ver_todo)
def eliminar_estudiante(request, pk):
    estudiante = get_object_or_404(_estudiantes_visibles_qs(request.user), pk=pk)

    if request.method == 'POST':
        estudiante.activo = False
        estudiante.save()
        messages.success(
            request,
            f'Estudiante {estudiante.nombres} {estudiante.apellidos} inactivado correctamente.'
        )
        return redirect('lista_estudiantes')

    return render(request, 'academico/eliminar_estudiante.html', {'estudiante': estudiante})


@login_required
def exportar_estudiantes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Estudiantes'

    encabezados = [
        'Código',
        'Tipo documento',
        'Documento',
        'Nombres',
        'Apellidos',
        'Genero',
        'Fecha nacimiento',
        'Grado',
        'Grupo',
        'Correo estudiante',
        'WhatsApp estudiante',
        'Acudiente',
        'Correo acudiente',
        'Teléfono acudiente',
        'WhatsApp acudiente',
        'Dirección',
        'Estado',
    ]

    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    estudiantes, filtros = _estudiantes_filtrados(request)
    if not filtros['filtros_aplicados']:
        estudiantes = _estudiantes_visibles_qs(request.user).order_by('apellidos', 'nombres')

    for estudiante in estudiantes:
        ws.append([
            estudiante.codigo,
            estudiante.get_tipo_documento_display(),
            estudiante.documento,
            estudiante.nombres,
            estudiante.apellidos,
            estudiante.get_genero_display(),
            estudiante.fecha_nacimiento.strftime('%Y-%m-%d'),
            estudiante.grupo.grado.nombre,
            estudiante.grupo.nombre,
            estudiante.correo or '',
            estudiante.whatsapp or '',
            estudiante.acudiente,
            estudiante.correo_acudiente or '',
            estudiante.telefono_acudiente,
            estudiante.whatsapp_acudiente or '',
            estudiante.direccion or '',
            'Activo' if estudiante.activo else 'Inactivo',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=estudiantes.xlsx'
    return response

@role_required(puede_ver_todo)
def importar_estudiantes(request):
    errores = []

    if request.method == 'POST':
        form = ImportarEstudiantesForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']

            try:
                wb = load_workbook(archivo)
                ws = wb.active

                total = 0

                for indice, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not fila or not fila[0]:
                            continue

                        tipo_documento = str(fila[0]).strip()
                        documento = _normalizar_numero(fila[1])
                        nombres = str(fila[2]).strip()
                        apellidos = str(fila[3]).strip()
                        genero_raw = str(fila[4]).strip().upper() if fila[4] else ''
                        genero = {
                            'M': 'M',
                            'H': 'M',
                            'HOMBRE': 'M',
                            'MASCULINO': 'M',
                            'F': 'F',
                            'MUJER': 'F',
                            'FEMENINO': 'F',
                        }.get(genero_raw, inferir_genero_por_nombre(nombres))
                        fecha_nacimiento = _parsear_fecha_excel(fila[5])
                        grado_nombre = str(fila[6]).strip() if len(fila) > 6 and fila[6] else ''
                        grupo_nombre = str(fila[7]).strip() if len(fila) > 7 and fila[7] else ''
                        grupo_id = fila[8] if len(fila) > 8 else ''
                        correo = str(fila[9]).strip() if len(fila) > 9 and fila[9] else ''
                        whatsapp = _normalizar_numero(fila[10]) if len(fila) > 10 and fila[10] else ''
                        acudiente = str(fila[11]).strip() if len(fila) > 11 and fila[11] else ''
                        correo_acudiente = str(fila[12]).strip() if len(fila) > 12 and fila[12] else ''
                        telefono_acudiente = _normalizar_numero(fila[13]) if len(fila) > 13 else ''
                        whatsapp_acudiente = _normalizar_numero(fila[14]) if len(fila) > 14 and fila[14] else ''
                        direccion = str(fila[15]).strip() if len(fila) > 15 and fila[15] else ''
                        activo_val = str(fila[16]).strip().lower() if len(fila) > 16 and fila[16] is not None else ''

                        if not all([
                            tipo_documento,
                            documento,
                            nombres,
                            apellidos,
                            genero_raw,
                            fecha_nacimiento,
                            grado_nombre,
                            grupo_nombre,
                            grupo_id,
                            correo,
                            whatsapp,
                            acudiente,
                            correo_acudiente,
                            telefono_acudiente,
                            whatsapp_acudiente,
                            direccion,
                            activo_val,
                        ]):
                            errores.append(f'Fila {indice}: faltan datos obligatorios.')
                            continue

                        try:
                            grupo = _resolver_grupo_por_grado_y_grupo(grado_nombre, grupo_nombre)
                        except Grupo.DoesNotExist:
                            errores.append(f'Fila {indice}: la combinacion grado "{grado_nombre}" y grupo "{grupo_nombre}" no existe.')
                            continue

                        if str(grupo.id) != str(grupo_id).strip():
                            errores.append(f'Fila {indice}: el grupo_id {grupo_id} no coincide con {grado_nombre} - {grupo_nombre}.')
                            continue

                        if not fecha_nacimiento:
                            errores.append(f'Fila {indice}: fecha de nacimiento inválida.')
                            continue

                        if len(documento) < 8:
                            errores.append(f'Fila {indice}: el documento debe tener minimo 8 digitos.')
                            continue

                        activo = activo_val in ['true', '1', 'si', 'activo']

                        if Estudiante.objects.filter(documento=documento).exists():
                            errores.append(f'Fila {indice}: el documento {documento} ya existe.')
                            continue

                        Estudiante.objects.create(
                            tipo_documento=tipo_documento,
                            documento=documento,
                            nombres=nombres,
                            apellidos=apellidos,
                            genero=genero if genero in {'M', 'F'} else inferir_genero_por_nombre(nombres),
                            fecha_nacimiento=fecha_nacimiento,
                            grupo=grupo,
                            correo=correo,
                            whatsapp=whatsapp,
                            acudiente=acudiente,
                            correo_acudiente=correo_acudiente,
                            telefono_acudiente=telefono_acudiente,
                            whatsapp_acudiente=whatsapp_acudiente,
                            direccion=direccion,
                            activo=activo,
                        )
                        total += 1

                    except Exception as e:
                        errores.append(f'Fila {indice}: error inesperado ({e}).')

                if total > 0:
                    messages.success(request, f'Se importaron {total} estudiantes correctamente.')

                if errores:
                    messages.warning(request, 'El archivo se procesó con algunas observaciones.')

            except Exception as e:
                messages.error(request, f'Error al importar archivo: {e}')
    else:
        form = ImportarEstudiantesForm()

    return render(request, 'academico/importar_estudiantes.html', {
        'form': form,
        'errores': errores
    })


@login_required
def mi_perfil_estudiante(request):
    if not es_estudiante(request.user):
        return redirect('lista_estudiantes')
    estudiante = _estudiante_sesion_o_404(request.user)
    return redirect('detalle_estudiante', pk=estudiante.pk)


@login_required
def detalle_estudiante(request, pk):
    estudiante = get_object_or_404(
        _estudiantes_visibles_qs(request.user),
        pk=pk
    )
    horarios = HorarioClase.objects.select_related(
        'carga_academica',
        'carga_academica__asignatura',
        'carga_academica__docente',
        'carga_academica__grupo',
    ).filter(
        carga_academica__grupo=estudiante.grupo,
        carga_academica__activo=True,
    ).order_by('dia_semana', 'hora_inicio', 'carga_academica__asignatura__nombre')
    back_url = _obtener_next_seguro(request)
    return render(request, 'academico/detalle_estudiante.html', {
        'estudiante': estudiante,
        'horarios': horarios,
        'back_url': back_url,
    })


@login_required
def horarios_academicos(request):
    anio_id = request.GET.get('anio', '').strip()
    grado_id = request.GET.get('grado', '').strip()
    grupo_id = request.GET.get('grupo', '').strip()
    asignatura_id = request.GET.get('asignatura', '').strip()
    dia_id = request.GET.get('dia', '').strip()
    cargas = cargas_visibles_para(request.user)
    estudiante_sesion = _estudiante_sesion_o_404(request.user) if es_estudiante(request.user) else None
    docente_sesion = request.user if es_docente(request.user) else None
    docente_id = str(docente_sesion.id) if docente_sesion else request.GET.get('docente', '').strip()

    anios = AnioLectivo.objects.filter(
        id__in=cargas.values_list('anio_lectivo_id', flat=True).distinct()
    ).order_by('-anio')

    if not anio_id:
        anio_activo = anios.filter(activo=True).first()
        if anio_activo:
            anio_id = str(anio_activo.id)

    if estudiante_sesion:
        grupo_id = str(estudiante_sesion.grupo_id)
        grado_id = str(estudiante_sesion.grupo.grado_id)
        docente_id = ''

    if anio_id:
        cargas = cargas.filter(anio_lectivo_id=anio_id)

    if grado_id:
        cargas = cargas.filter(grupo__grado_id=grado_id)

    if grupo_id:
        cargas = cargas.filter(grupo_id=grupo_id)

    if asignatura_id:
        cargas = cargas.filter(asignatura_id=asignatura_id)

    if docente_id:
        cargas = cargas.filter(docente_id=docente_id)

    cargas_filtradas = cargas

    horarios = HorarioClase.objects.select_related(
        'carga_academica',
        'carga_academica__grupo',
        'carga_academica__grupo__grado',
        'carga_academica__asignatura',
        'carga_academica__anio_lectivo',
        'carga_academica__docente',
    ).filter(
        carga_academica__activo=True,
        carga_academica__in=cargas,
    )

    if dia_id:
        horarios = horarios.filter(dia_semana=dia_id)

    grupos = grupos_visibles_para(request.user).filter(
        id__in=cargas_filtradas.values_list('grupo_id', flat=True).distinct()
    ).distinct().order_by('grado__nombre', 'nombre')
    grados = Grado.objects.filter(
        id__in=grupos.values_list('grado_id', flat=True).distinct()
    ).order_by('nombre')
    asignaturas = Asignatura.objects.filter(
        id__in=cargas_filtradas.values_list('asignatura_id', flat=True).distinct()
    ).order_by('nombre')
    docentes = Usuario.objects.filter(
        id__in=cargas_filtradas.values_list('docente_id', flat=True).distinct()
    ).order_by('first_name', 'last_name', 'username')

    horarios = horarios.order_by(
        'dia_semana', 'hora_inicio',
        'carga_academica__grupo__grado__nombre',
        'carga_academica__grupo__nombre',
    )

    horarios_por_dia = {}
    for horario in horarios:
        horarios_por_dia.setdefault(horario.get_dia_semana_display(), []).append(horario)

    dias_ordenados = list(horarios_por_dia.items())
    filas_horario = [
        dias_ordenados[indice:indice + 2]
        for indice in range(0, len(dias_ordenados), 2)
    ]

    cargas_con_horario_ids = HorarioClase.objects.filter(
        carga_academica__in=cargas_filtradas
    ).values_list('carga_academica_id', flat=True).distinct()
    cargas_sin_horario = cargas_filtradas.exclude(
        id__in=cargas_con_horario_ids
    ).select_related(
        'docente',
        'grupo',
        'grupo__grado',
        'asignatura',
        'anio_lectivo',
    ).order_by(
        'grupo__grado__nombre',
        'grupo__nombre',
        'asignatura__nombre',
    )
    if estudiante_sesion:
        cargas_sin_horario = cargas_sin_horario.none()

    total_bloques = horarios.count()
    resumen = {
        'bloques': total_bloques,
        'grupos': horarios.values('carga_academica__grupo_id').distinct().count(),
        'docentes': horarios.values('carga_academica__docente_id').distinct().count(),
        'aulas': horarios.exclude(aula__isnull=True).exclude(aula='').values('aula').distinct().count(),
        'sin_horario': cargas_sin_horario.count(),
    }

    return render(request, 'academico/horarios_academicos.html', {
        'horarios_por_dia': horarios_por_dia,
        'filas_horario': filas_horario,
        'anios': anios,
        'grados': grados,
        'grupos': grupos,
        'asignaturas': asignaturas,
        'docentes': docentes,
        'docente_sesion': docente_sesion,
        'estudiante_sesion': estudiante_sesion,
        'anio_id': anio_id,
        'grado_id': grado_id,
        'grupo_id': grupo_id,
        'asignatura_id': asignatura_id,
        'dia_id': dia_id,
        'docente_id': docente_id,
        'total_bloques': total_bloques,
        'resumen': resumen,
        'dias_semana': HorarioClase.DIA_CHOICES,
        'cargas_sin_horario': cargas_sin_horario,
    })


@login_required
def gestion_seguimiento(request):
    busqueda = request.GET.get('busqueda', '').strip()
    grado_id = request.GET.get('grado', '').strip()
    grupo_id = request.GET.get('grupo', '').strip()
    periodo = request.GET.get('periodo', 'semanal').strip()
    fecha_inicio_q = request.GET.get('fecha_inicio', '').strip()
    fecha_fin_q = request.GET.get('fecha_fin', '').strip()
    estado_reporte = request.GET.get('estado', '').strip().lower()
    estado_alerta = request.GET.get('estado_alerta', '').strip()
    nivel_alerta = request.GET.get('nivel_alerta', '').strip()
    filtrar = request.GET.get('filtrar', '').strip() == '1'

    estudiantes_visibles = filtrar_estudiantes_visibles(
        Estudiante.objects.select_related('grupo', 'grupo__grado').filter(activo=True),
        request.user
    )
    grupos = Grupo.objects.select_related('grado').filter(
        id__in=estudiantes_visibles.values_list('grupo_id', flat=True).distinct()
    ).order_by('grado__nombre', 'nombre')
    grados = Grado.objects.filter(grupos__in=grupos).distinct().order_by('nombre')
    todos_grupos = grupos

    if grado_id and grupo_id and not grupos.filter(id=grupo_id, grado_id=grado_id).exists():
        grupo_id = ''

    periodo = periodo if periodo in PERIODOS_REPORTE else 'semanal'
    fecha_inicio = _parsear_fecha_query(fecha_inicio_q)
    fecha_fin = _parsear_fecha_query(fecha_fin_q)
    if periodo == 'semanal':
        hoy = datetime.now().date()
        lunes_actual = hoy - timedelta(days=hoy.weekday())
        viernes_actual = lunes_actual + timedelta(days=4)
        fecha_inicio = fecha_inicio or lunes_actual
        fecha_fin = fecha_fin or viernes_actual
        if fecha_fin < fecha_inicio:
            fecha_fin = fecha_inicio + timedelta(days=4)
    else:
        fecha_inicio = fecha_inicio or None
        fecha_fin = fecha_fin or None
    reportes_base = ReporteAcudiente.objects.filter(estudiante__in=estudiantes_visibles)
    enviados_visibles = estudiantes_visibles.annotate(
        tiene_reporte_enviado_global=Exists(
            ReporteAcudiente.objects.filter(
                estudiante_id=OuterRef('pk'),
                estado='ENVIADO',
            )
        )
    )
    alertas_base = filtrar_alertas_visibles(
        AlertaTemprana.objects.select_related(
            'estudiante',
            'estudiante__grupo',
            'estudiante__grupo__grado',
            'tipo_alerta',
            'configuracion',
        ),
        request.user
    )
    alertas_activas = alertas_base.filter(estado='ACTIVA')

    estudiantes = estudiantes_visibles.none()
    alertas = alertas_base.none()
    page_obj = Paginator(estudiantes_visibles.none(), 10).get_page(1)
    alerta_page_obj = Paginator(alertas_base.none(), 8).get_page(1)
    total_estudiantes_filtrados = 0
    total_pendientes_filtrados = 0
    total_alertas_filtradas = 0
    total_activas_filtradas = 0
    total_criticas_filtradas = 0
    con_correo = 0

    if filtrar:
        estudiantes = estudiantes_visibles

        if busqueda:
            estudiantes = estudiantes.filter(
                Q(codigo__icontains=busqueda) |
                Q(documento__icontains=busqueda) |
                Q(nombres__icontains=busqueda) |
                Q(apellidos__icontains=busqueda) |
                Q(acudiente__icontains=busqueda) |
                Q(correo_acudiente__icontains=busqueda)
            )

        if grado_id:
            estudiantes = estudiantes.filter(grupo__grado_id=grado_id)
            grupos = grupos.filter(grado_id=grado_id)

        if grupo_id:
            estudiantes = estudiantes.filter(grupo_id=grupo_id)

        if estado_alerta:
            estudiantes = estudiantes.filter(alertas__estado=estado_alerta)

        if nivel_alerta:
            estudiantes = estudiantes.filter(alertas__nivel=nivel_alerta)

        reporte_periodo_subquery = ReporteAcudiente.objects.filter(
            estudiante_id=OuterRef('pk'),
            periodo=periodo,
            estado='ENVIADO',
        )
        estudiantes = estudiantes.annotate(
            tiene_reporte_enviado=Exists(reporte_periodo_subquery),
            total_alertas_activas=Count('alertas', filter=Q(alertas__estado='ACTIVA'), distinct=True),
            total_alertas_criticas=Count('alertas', filter=Q(alertas__estado='ACTIVA', alertas__nivel='CRITICO'), distinct=True),
        ).distinct()

        if estado_reporte == 'enviado':
            estudiantes = estudiantes.filter(tiene_reporte_enviado=True)
        elif estado_reporte == 'pendiente':
            estudiantes = estudiantes.filter(tiene_reporte_enviado=False)

        alertas = alertas_base

        if busqueda:
            alertas = alertas.filter(
                Q(estudiante__nombres__icontains=busqueda) |
                Q(estudiante__apellidos__icontains=busqueda) |
                Q(estudiante__codigo__icontains=busqueda) |
                Q(estudiante__documento__icontains=busqueda)
            )

        if grado_id:
            alertas = alertas.filter(estudiante__grupo__grado_id=grado_id)

        if grupo_id:
            alertas = alertas.filter(estudiante__grupo_id=grupo_id)

        if estado_alerta:
            alertas = alertas.filter(estado=estado_alerta)

        if nivel_alerta:
            alertas = alertas.filter(nivel=nivel_alerta)

        alertas = alertas.filter(estudiante__in=estudiantes.values('pk')).order_by('-fecha_generacion')

        total_estudiantes_filtrados = estudiantes.count()
        total_pendientes_filtrados = estudiantes.filter(tiene_reporte_enviado=False).count()
        total_alertas_filtradas = alertas.count()
        total_activas_filtradas = alertas.filter(estado='ACTIVA').count()
        total_criticas_filtradas = alertas.filter(estado='ACTIVA', nivel='CRITICO').count()
        con_correo = estudiantes.exclude(correo_acudiente__isnull=True).exclude(correo_acudiente='').count()

        page_obj = Paginator(estudiantes.order_by('apellidos', 'nombres'), 10).get_page(request.GET.get('page'))
        alerta_page_obj = Paginator(alertas, 8).get_page(request.GET.get('alerta_page'))
        for alerta in alerta_page_obj.object_list:
            alerta.descripcion_visible = construir_descripcion_actual_alerta(alerta)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    query_params.pop('alerta_page', None)
    querystring = query_params.urlencode()

    return render(request, 'academico/gestion_seguimiento.html', {
        'page_obj': page_obj,
        'alerta_page_obj': alerta_page_obj,
        'busqueda': busqueda,
        'grado_id': grado_id,
        'grupo_id': grupo_id,
        'periodo': periodo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'estado': estado_reporte,
        'estado_alerta': estado_alerta,
        'nivel_alerta': nivel_alerta,
        'filtrar': filtrar,
        'filtros_aplicados': filtrar,
        'grados': grados,
        'grupos': grupos,
        'todos_grupos': todos_grupos,
        'querystring': querystring,
        'total_estudiantes': total_estudiantes_filtrados if filtrar else estudiantes_visibles.count(),
        'con_correo': con_correo,
        'total_reportes_enviados': reportes_base.filter(estado='ENVIADO').count(),
        'total_reportes_pendientes': total_pendientes_filtrados if filtrar else enviados_visibles.filter(tiene_reporte_enviado_global=False).count(),
        'total_alertas_activas': alertas_activas.count(),
        'alertas_filtradas': total_alertas_filtradas,
        'alertas_activas_filtradas': total_activas_filtradas,
        'alertas_criticas_filtradas': total_criticas_filtradas,
        'estado_correo': obtener_estado_correo(),
        'page_numbers': _page_window(page_obj),
        'page_param': 'page',
        'alerta_page_numbers': _page_window(alerta_page_obj),
        'alerta_page_param': 'alerta_page',
    })


@login_required
def gestion_reportes(request):
    return gestion_seguimiento(request)


@login_required
def mi_seguimiento_estudiante(request):
    if not es_estudiante(request.user):
        return redirect('gestion_seguimiento')
    estudiante = _estudiante_sesion_o_404(request.user)
    periodo = request.GET.get('periodo', 'semanal')
    query = request.GET.urlencode()
    url = reverse('reporte_estudiante', args=[estudiante.pk])
    return redirect(f'{url}?{query}' if query else f'{url}?periodo={periodo}')


@login_required
def reporte_estudiante(request, pk):
    estudiante = get_object_or_404(
        _estudiantes_visibles_qs(request.user),
        pk=pk
    )
    periodo = request.GET.get('periodo', 'semanal')
    fecha_inicio = _parsear_fecha_query(request.GET.get('fecha_inicio', '').strip())
    fecha_fin = _parsear_fecha_query(request.GET.get('fecha_fin', '').strip())
    datos = construir_datos_reporte_estudiante(estudiante, periodo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    return render(request, 'academico/reporte_estudiante.html', datos)


@login_required
def descargar_reporte_estudiante(request, pk, periodo):
    estudiante = get_object_or_404(
        _estudiantes_visibles_qs(request.user),
        pk=pk
    )
    fecha_inicio = _parsear_fecha_query(request.GET.get('fecha_inicio', '').strip())
    fecha_fin = _parsear_fecha_query(request.GET.get('fecha_fin', '').strip())
    datos = construir_datos_reporte_estudiante(estudiante, periodo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    contenido = generar_excel_reporte_estudiante(datos)

    response = HttpResponse(
        contenido,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{datos["periodo"]}_{estudiante.codigo}.xlsx"'
    registrar_reporte(
        datos,
        'DESCARGADO',
        destinatario=estudiante.correo_acudiente or '',
        enviado_por=request.user,
        asunto=f'Descarga reporte {datos["periodo"]} de {estudiante.nombres} {estudiante.apellidos}',
    )
    return response


@role_required(puede_gestionar_docencia)
def enviar_reporte_estudiante(request, pk, periodo):
    estudiante = get_object_or_404(
        _estudiantes_visibles_qs(request.user),
        pk=pk
    )
    fecha_inicio = _parsear_fecha_query(request.GET.get('fecha_inicio', '').strip())
    fecha_fin = _parsear_fecha_query(request.GET.get('fecha_fin', '').strip())
    datos = construir_datos_reporte_estudiante(estudiante, periodo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

    try:
        destinatario = enviar_reporte_estudiante_por_correo(datos, enviado_por=request.user)
        messages.success(request, f'Reporte {datos["periodo_label"].lower()} enviado correctamente a {destinatario}.')
    except ValueError as exc:
        messages.error(request, str(exc))
    except Exception as exc:
        messages.error(request, f'No fue posible enviar el correo: {exc}')

    redirect_url = f'/academico/estudiantes/{estudiante.pk}/reporte/?periodo={datos["periodo"]}'
    if datos.get('fecha_inicio'):
        redirect_url += f'&fecha_inicio={datos["fecha_inicio"].strftime("%Y-%m-%d")}'
    if datos.get('fecha_fin'):
        redirect_url += f'&fecha_fin={datos["fecha_fin"].strftime("%Y-%m-%d")}'
    return redirect(redirect_url)


@role_required(puede_ver_todo)
def enviar_reportes_masivos(request):
    if request.method != 'POST':
        return redirect('gestion_reportes')

    mutable_get = request.GET.copy()
    mutable_get['filtrar'] = '1'
    request.GET = mutable_get
    estudiantes, filtros = _estudiantes_reportes_filtrados(request)
    estudiantes = estudiantes.exclude(correo_acudiente__isnull=True).exclude(correo_acudiente='')
    fecha_inicio = _parsear_fecha_query(request.GET.get('fecha_inicio', '').strip())
    fecha_fin = _parsear_fecha_query(request.GET.get('fecha_fin', '').strip())

    enviados, errores = enviar_reportes_estudiantes(
        estudiantes,
        filtros['periodo'],
        enviado_por=request.user,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    if enviados:
        messages.success(
            request,
            f'Se enviaron {len(enviados)} reportes {filtros["periodo"]} correctamente.'
        )
    if errores:
        messages.warning(
            request,
            f'No fue posible enviar {len(errores)} reportes. Revisa los correos configurados y la salida SMTP.'
        )

    return redirect(f'/academico/seguimiento/?{request.GET.urlencode()}')


@role_required(puede_ver_todo)
def probar_correo_reportes(request):
    if request.method != 'POST':
        return redirect('gestion_seguimiento')

    estado = obtener_estado_correo()
    if not estado['configurado']:
        messages.warning(request, estado['mensaje'])
        return redirect('gestion_seguimiento')

    try:
        probar_conexion_correo()
        messages.success(request, 'La conexion SMTP fue validada correctamente desde la aplicacion.')
    except Exception as exc:
        messages.error(request, f'La conexion SMTP fallo: {exc}')

    return redirect('gestion_seguimiento')


@role_required(puede_gestionar_catalogos)
def gestion_academica(request):
    resumen = []
    for slug, config in CATALOGOS.items():
        resumen.append({
            'slug': slug,
            'titulo': config['titulo'],
            'total': config['queryset']().count(),
        })

    return render(request, 'academico/gestion_academica.html', {'resumen': resumen})


@role_required(puede_gestionar_catalogos)
def lista_catalogo(request, catalogo):
    config = _obtener_catalogo(catalogo)
    objetos = config['queryset']()
    filas = [
        {
            'objeto': objeto,
            'valores': _formatear_fila(objeto, config['columns']),
        }
        for objeto in objetos
    ]

    return render(request, 'academico/catalogo_lista.html', {
        'catalogo': catalogo,
        'titulo': config['titulo'],
        'singular': config['singular'],
        'columns': config['columns'],
        'filas': filas,
    })


@role_required(puede_gestionar_catalogos)
def crear_catalogo(request, catalogo):
    config = _obtener_catalogo(catalogo)
    form_class = config['form']

    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'{config["singular"].capitalize()} creado correctamente.')
            return redirect('lista_catalogo', catalogo=catalogo)
    else:
        form = form_class()

    return render(request, 'academico/catalogo_form.html', {
        'catalogo': catalogo,
        'titulo': f'Crear {config["singular"]}',
        'form': form,
        'boton': 'Guardar',
    })


@role_required(puede_gestionar_catalogos)
def editar_catalogo(request, catalogo, pk):
    config = _obtener_catalogo(catalogo)
    objeto = get_object_or_404(config['model'], pk=pk)
    form_class = config['form']

    if request.method == 'POST':
        form = form_class(request.POST, instance=objeto)
        if form.is_valid():
            form.save()
            messages.success(request, f'{config["singular"].capitalize()} actualizado correctamente.')
            return redirect('lista_catalogo', catalogo=catalogo)
    else:
        form = form_class(instance=objeto)

    return render(request, 'academico/catalogo_form.html', {
        'catalogo': catalogo,
        'titulo': f'Editar {config["singular"]}',
        'form': form,
        'boton': 'Guardar cambios',
        'objeto': objeto,
    })


@role_required(puede_gestionar_catalogos)
def eliminar_catalogo(request, catalogo, pk):
    config = _obtener_catalogo(catalogo)
    objeto = get_object_or_404(config['model'], pk=pk)

    if request.method == 'POST':
        objeto.delete()
        messages.success(request, f'{config["singular"].capitalize()} eliminado correctamente.')
        return redirect('lista_catalogo', catalogo=catalogo)

    return render(request, 'academico/catalogo_confirmar_eliminar.html', {
        'catalogo': catalogo,
        'titulo': f'Eliminar {config["singular"]}',
        'objeto': objeto,
    })
