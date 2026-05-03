from collections import OrderedDict
import base64
from datetime import datetime, timedelta
import json
from io import BytesIO
from urllib import request as urllib_request
from urllib.error import HTTPError, URLError

from django.conf import settings
from django.core.mail import EmailMultiAlternatives, get_connection
from django.template.loader import render_to_string
from django.utils.timezone import localdate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from academico.models import ReporteAcudiente
from alertas.models import AlertaTemprana
from asistencia.models import Asistencia
from evaluacion.models import Calificacion
from evaluacion.utils import calcular_promedio_dimensionado


PERIODOS_REPORTE = {
    'semanal': {
        'label': 'Semanal',
        'dias': 4,
    },
    'mensual': {
        'label': 'Mensual',
        'dias': 29,
    },
}


def _usa_brevo_api():
    return getattr(settings, 'EMAIL_TRANSPORT', 'smtp') == 'brevo_api' and bool(getattr(settings, 'BREVO_API_KEY', ''))


def _parsear_remitente(valor):
    valor = (valor or '').strip()
    if '<' in valor and '>' in valor:
        nombre, correo = valor.split('<', 1)
        return nombre.strip().strip('"'), correo.replace('>', '').strip()
    return '', valor


def _enviar_correo_brevo_api(*, asunto, cuerpo_texto, cuerpo_html, destinatario, nombre_destinatario, nombre_archivo, contenido):
    remitente_nombre, remitente_correo = _parsear_remitente(settings.DEFAULT_FROM_EMAIL)
    if not remitente_correo:
        raise ValueError('DEFAULT_FROM_EMAIL no esta configurado correctamente para Brevo API.')

    payload = {
        'sender': {
            'name': remitente_nombre or 'Plataforma Academica',
            'email': remitente_correo,
        },
        'to': [
            {
                'email': destinatario,
                'name': nombre_destinatario or destinatario,
            }
        ],
        'subject': asunto,
        'htmlContent': cuerpo_html,
        'textContent': cuerpo_texto,
        'attachment': [
            {
                'name': nombre_archivo,
                'content': base64.b64encode(contenido).decode('ascii'),
            }
        ],
    }

    req = urllib_request.Request(
        settings.BREVO_API_URL,
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'accept': 'application/json',
            'api-key': settings.BREVO_API_KEY,
            'content-type': 'application/json',
        },
        method='POST',
    )

    try:
        with urllib_request.urlopen(req, timeout=settings.EMAIL_TIMEOUT) as response:
            status_code = getattr(response, 'status', None) or response.getcode()
            if status_code >= 400:
                raise ValueError(f'Brevo API respondio con estado {status_code}.')
    except HTTPError as exc:
        detalle = exc.read().decode('utf-8', errors='ignore')
        raise ValueError(f'Brevo API devolvio HTTP {exc.code}: {detalle or exc.reason}') from exc
    except URLError as exc:
        raise ValueError(f'No fue posible conectar con Brevo API: {exc.reason}') from exc


def _preparar_hoja_excel(hoja, titulo, encabezados):
    hoja.title = titulo
    hoja.append(encabezados)
    fill = PatternFill('solid', fgColor='1F2937')
    for celda in hoja[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = fill


def _ajustar_columnas(hoja):
    for columna in hoja.columns:
        longitud = max(len(str(celda.value or '')) for celda in columna)
        hoja.column_dimensions[get_column_letter(columna[0].column)].width = min(longitud + 2, 45)


def obtener_periodo_reporte(periodo):
    return PERIODOS_REPORTE.get(periodo, PERIODOS_REPORTE['semanal'])


def _resolver_rango_reporte(periodo, fecha_inicio=None, fecha_fin=None):
    configuracion = obtener_periodo_reporte(periodo)
    hoy = localdate()

    if fecha_inicio and isinstance(fecha_inicio, str):
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin and isinstance(fecha_fin, str):
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

    if periodo == 'semanal':
        lunes_actual = hoy - timedelta(days=hoy.weekday())
        viernes_actual = lunes_actual + timedelta(days=4)
        fecha_inicio = fecha_inicio or lunes_actual
        fecha_fin = fecha_fin or viernes_actual

        if fecha_inicio.weekday() != 0:
            fecha_inicio = fecha_inicio - timedelta(days=fecha_inicio.weekday())
        if fecha_fin.weekday() != 4:
            fecha_fin = fecha_inicio + timedelta(days=4)
        if fecha_fin < fecha_inicio:
            fecha_fin = fecha_inicio + timedelta(days=4)
        return fecha_inicio, fecha_fin, configuracion

    fecha_fin = fecha_fin or hoy
    fecha_inicio = fecha_inicio or (fecha_fin - timedelta(days=configuracion['dias']))
    if fecha_fin < fecha_inicio:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio
    return fecha_inicio, fecha_fin, configuracion


def _agrupar_seguimiento_diario(asistencias, calificaciones):
    seguimiento = OrderedDict()

    for asistencia in sorted(asistencias, key=lambda item: (item.fecha, item.carga_academica.asignatura.nombre)):
        fecha = asistencia.fecha
        fecha_item = seguimiento.setdefault(fecha, OrderedDict())
        asignatura = asistencia.carga_academica.asignatura.nombre
        materia = fecha_item.setdefault(asignatura, {
            'asignatura': asignatura,
            'asistencia': None,
            'calificaciones': [],
        })
        materia['asistencia'] = asistencia

    for calificacion in sorted(
        calificaciones,
        key=lambda item: (item.actividad.fecha, item.actividad.carga_academica.asignatura.nombre, item.actividad.nombre)
    ):
        fecha = calificacion.actividad.fecha
        fecha_item = seguimiento.setdefault(fecha, OrderedDict())
        asignatura = calificacion.actividad.carga_academica.asignatura.nombre
        materia = fecha_item.setdefault(asignatura, {
            'asignatura': asignatura,
            'asistencia': None,
            'calificaciones': [],
        })
        materia['calificaciones'].append(calificacion)

    dias = []
    for fecha, materias in reversed(seguimiento.items()):
        materias_ordenadas = list(materias.values())
        dias.append({
            'fecha': fecha,
            'materias': materias_ordenadas,
            'total_materias': len(materias_ordenadas),
            'materias_con_nota': sum(1 for materia in materias_ordenadas if materia['calificaciones']),
            'materias_con_asistencia': sum(1 for materia in materias_ordenadas if materia['asistencia']),
        })

    return dias


def _completar_dias_habiles(dias, fecha_inicio, fecha_fin):
    dias_por_fecha = {dia['fecha']: dia for dia in dias}
    dias_completos = []
    cursor = fecha_inicio

    while cursor <= fecha_fin:
        if cursor.weekday() < 5:
            dias_completos.append(
                dias_por_fecha.get(cursor, {
                    'fecha': cursor,
                    'materias': [],
                    'total_materias': 0,
                    'materias_con_nota': 0,
                    'materias_con_asistencia': 0,
                })
            )
        cursor += timedelta(days=1)

    return dias_completos


def _construir_horario_por_dia(horarios):
    dias = OrderedDict([
        ('Lunes', []),
        ('Martes', []),
        ('Miercoles', []),
        ('Jueves', []),
        ('Viernes', []),
    ])

    for carga in horarios:
        for horario in carga.horarios.all():
            dia = horario.get_dia_semana_display()
            dias.setdefault(dia, []).append({
                'asignatura': carga.asignatura.nombre,
                'docente': carga.docente.get_full_name() or carga.docente.username,
                'hora_inicio': horario.hora_inicio,
                'hora_fin': horario.hora_fin,
                'aula': horario.aula or '-',
            })

    dias_ordenados = []
    for dia, bloques in dias.items():
        bloques_ordenados = sorted(bloques, key=lambda item: (item['hora_inicio'], item['asignatura']))
        if bloques_ordenados:
            dias_ordenados.append({
                'dia': dia,
                'bloques': bloques_ordenados,
            })

    return dias_ordenados


def construir_datos_reporte_estudiante(estudiante, periodo, fecha_inicio=None, fecha_fin=None):
    fecha_inicio, fecha_fin, configuracion = _resolver_rango_reporte(periodo, fecha_inicio, fecha_fin)

    asistencias = list(
        Asistencia.objects.filter(
            estudiante=estudiante,
            fecha__range=(fecha_inicio, fecha_fin)
        ).select_related(
            'carga_academica',
            'carga_academica__asignatura',
            'carga_academica__grupo',
        ).order_by('-fecha', 'carga_academica__asignatura__nombre')
    )

    calificaciones = list(
        Calificacion.objects.filter(
            estudiante=estudiante,
            actividad__fecha__range=(fecha_inicio, fecha_fin)
        ).select_related(
            'actividad',
            'actividad__carga_academica',
            'actividad__carga_academica__asignatura',
            'actividad__carga_academica__grupo',
            'actividad__periodo',
        ).order_by('-actividad__fecha', 'actividad__nombre')
    )

    alertas = list(
        AlertaTemprana.objects.filter(
            estudiante=estudiante,
            fecha_generacion__date__range=(fecha_inicio, fecha_fin)
        ).select_related(
            'tipo_alerta',
            'configuracion',
        ).prefetch_related('seguimientos').order_by('-fecha_generacion')
    )

    total_presentes = sum(1 for item in asistencias if item.estado == 'P')
    total_ausentes = sum(1 for item in asistencias if item.estado == 'A')
    total_tardes = sum(1 for item in asistencias if item.estado == 'T')
    total_justificadas = sum(1 for item in asistencias if item.estado == 'J')
    total_registros = len(asistencias)
    porcentaje_asistencia = round((total_presentes / total_registros) * 100, 1) if total_registros else 0

    total_calificaciones = len(calificaciones)
    promedio_dimensionado, _ = calcular_promedio_dimensionado(calificaciones)
    promedio_simple = float(promedio_dimensionado) if promedio_dimensionado is not None else 0
    notas_bajas = sum(1 for item in calificaciones if item.nota < 3)
    alertas_activas = estudiante.alertas.filter(estado='ACTIVA').count()
    horarios = list(
        estudiante.grupo.cargas_academicas.filter(activo=True)
        .select_related('asignatura', 'docente')
        .prefetch_related('horarios')
        .order_by('asignatura__nombre')
    )
    horario_por_dia = _construir_horario_por_dia(horarios)
    seguimiento_diario = _agrupar_seguimiento_diario(asistencias, calificaciones)
    if periodo == 'semanal':
        seguimiento_diario = _completar_dias_habiles(seguimiento_diario, fecha_inicio, fecha_fin)

    return {
        'periodo': periodo if periodo in PERIODOS_REPORTE else 'semanal',
        'periodo_label': configuracion['label'],
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'semana_personalizada': periodo == 'semanal',
        'estudiante': estudiante,
        'horarios': horarios,
        'horario_por_dia': horario_por_dia,
        'asistencias': asistencias,
        'calificaciones': calificaciones,
        'seguimiento_diario': seguimiento_diario,
        'alertas': alertas,
        'resumen': {
            'total_presentes': total_presentes,
            'total_ausentes': total_ausentes,
            'total_tardes': total_tardes,
            'total_justificadas': total_justificadas,
            'total_registros': total_registros,
            'porcentaje_asistencia': porcentaje_asistencia,
            'total_calificaciones': total_calificaciones,
            'promedio_simple': promedio_simple,
            'notas_bajas': notas_bajas,
            'alertas_periodo': len(alertas),
            'alertas_activas': alertas_activas,
        }
    }


def generar_excel_reporte_estudiante(datos):
    estudiante = datos['estudiante']
    resumen = datos['resumen']

    libro = Workbook()

    hoja_resumen = libro.active
    _preparar_hoja_excel(hoja_resumen, 'Resumen', ['Campo', 'Valor'])
    for fila in [
        ('Estudiante', f'{estudiante.apellidos} {estudiante.nombres}'),
        ('Codigo', estudiante.codigo),
        ('Grupo', str(estudiante.grupo)),
        ('Periodo reporte', datos['periodo_label']),
        ('Fecha inicio', datos['fecha_inicio'].strftime('%Y-%m-%d')),
        ('Fecha fin', datos['fecha_fin'].strftime('%Y-%m-%d')),
        ('Presentes', resumen['total_presentes']),
        ('Ausentes', resumen['total_ausentes']),
        ('Tardes', resumen['total_tardes']),
        ('Justificadas', resumen['total_justificadas']),
        ('Porcentaje asistencia', resumen['porcentaje_asistencia']),
        ('Calificaciones registradas', resumen['total_calificaciones']),
        ('Promedio simple', resumen['promedio_simple']),
        ('Notas menores a 3.0', resumen['notas_bajas']),
        ('Alertas en el periodo', resumen['alertas_periodo']),
        ('Alertas activas', resumen['alertas_activas']),
    ]:
        hoja_resumen.append(fila)
    _ajustar_columnas(hoja_resumen)

    hoja_horario = libro.create_sheet('Horario')
    _preparar_hoja_excel(hoja_horario, 'Horario', [
        'Asignatura', 'Docente', 'Dia', 'Hora inicio', 'Hora fin', 'Aula'
    ])
    for carga in datos['horarios']:
        for horario in carga.horarios.all():
            hoja_horario.append([
                carga.asignatura.nombre,
                carga.docente.get_full_name() or carga.docente.username,
                horario.get_dia_semana_display(),
                horario.hora_inicio.strftime('%H:%M'),
                horario.hora_fin.strftime('%H:%M'),
                horario.aula or '',
            ])
    _ajustar_columnas(hoja_horario)

    hoja_asistencia = libro.create_sheet('Asistencias')
    _preparar_hoja_excel(hoja_asistencia, 'Asistencias', [
        'Fecha', 'Asignatura', 'Grupo', 'Estado', 'Observacion'
    ])
    for asistencia in datos['asistencias']:
        hoja_asistencia.append([
            asistencia.fecha.strftime('%Y-%m-%d'),
            asistencia.carga_academica.asignatura.nombre,
            str(asistencia.carga_academica.grupo),
            asistencia.get_estado_display(),
            asistencia.observacion or '',
        ])
    _ajustar_columnas(hoja_asistencia)

    hoja_calificaciones = libro.create_sheet('Calificaciones')
    _preparar_hoja_excel(hoja_calificaciones, 'Calificaciones', [
        'Fecha', 'Asignatura', 'Actividad', 'Periodo', 'Porcentaje', 'Nota', 'Observacion'
    ])
    for calificacion in datos['calificaciones']:
        hoja_calificaciones.append([
            calificacion.actividad.fecha.strftime('%Y-%m-%d'),
            calificacion.actividad.carga_academica.asignatura.nombre,
            calificacion.actividad.nombre,
            str(calificacion.actividad.periodo),
            float(calificacion.actividad.porcentaje),
            float(calificacion.nota),
            calificacion.observacion or '',
        ])
    _ajustar_columnas(hoja_calificaciones)

    hoja_seguimiento = libro.create_sheet('Seguimiento diario')
    _preparar_hoja_excel(hoja_seguimiento, 'Seguimiento diario', [
        'Fecha', 'Asignatura', 'Estado asistencia', 'Observacion asistencia',
        'Actividad', 'Periodo academico', 'Porcentaje', 'Nota', 'Observacion nota'
    ])
    for dia in datos['seguimiento_diario']:
        for materia in dia['materias']:
            asistencia = materia['asistencia']
            calificaciones = materia['calificaciones'] or [None]
            for calificacion in calificaciones:
                hoja_seguimiento.append([
                    dia['fecha'].strftime('%Y-%m-%d'),
                    materia['asignatura'],
                    asistencia.get_estado_display() if asistencia else 'Sin registro',
                    asistencia.observacion or '' if asistencia else '',
                    calificacion.actividad.nombre if calificacion else 'Sin actividad registrada',
                    str(calificacion.actividad.periodo) if calificacion else '',
                    float(calificacion.actividad.porcentaje) if calificacion else '',
                    float(calificacion.nota) if calificacion else '',
                    calificacion.observacion or '' if calificacion else '',
                ])
    _ajustar_columnas(hoja_seguimiento)

    hoja_alertas = libro.create_sheet('Alertas')
    _preparar_hoja_excel(hoja_alertas, 'Alertas', [
        'Fecha', 'Tipo', 'Nivel', 'Estado', 'Descripcion', 'Ultimo seguimiento'
    ])
    for alerta in datos['alertas']:
        seguimiento = alerta.seguimientos.first()
        hoja_alertas.append([
            alerta.fecha_generacion.strftime('%Y-%m-%d %H:%M'),
            alerta.tipo_alerta.nombre if alerta.tipo_alerta else '',
            alerta.nivel,
            alerta.estado,
            alerta.descripcion,
            seguimiento.descripcion if seguimiento else '',
        ])
    _ajustar_columnas(hoja_alertas)

    output = BytesIO()
    libro.save(output)
    return output.getvalue()


def obtener_estado_correo():
    if _usa_brevo_api():
        if not settings.DEFAULT_FROM_EMAIL:
            return {
                'backend': 'brevo_api',
                'configurado': False,
                'mensaje': 'Falta DEFAULT_FROM_EMAIL para el envio por Brevo API.'
            }
        return {
            'backend': 'brevo_api',
            'configurado': True,
            'mensaje': 'Configuracion lista para envio por Brevo API.'
        }

    backend = getattr(settings, 'EMAIL_BACKEND', '')
    if backend == 'django.core.mail.backends.console.EmailBackend':
        return {
            'backend': backend,
            'configurado': False,
            'mensaje': 'El backend actual es de consola. Los correos se imprimen, no se envian realmente.'
        }

    if backend == 'django.core.mail.backends.locmem.EmailBackend':
        return {
            'backend': backend,
            'configurado': True,
            'mensaje': 'El backend en memoria esta listo para pruebas de envio de correo.'
        }

    campos_faltantes = [
        campo for campo, valor in {
            'EMAIL_HOST': settings.EMAIL_HOST,
            'EMAIL_HOST_USER': settings.EMAIL_HOST_USER,
            'EMAIL_HOST_PASSWORD': settings.EMAIL_HOST_PASSWORD,
            'DEFAULT_FROM_EMAIL': settings.DEFAULT_FROM_EMAIL,
        }.items() if not valor
    ]

    if campos_faltantes:
        return {
            'backend': backend,
            'configurado': False,
            'mensaje': f'Faltan variables SMTP: {", ".join(campos_faltantes)}.'
        }

    return {
        'backend': backend,
        'configurado': True,
        'mensaje': f'Configuracion SMTP lista con backend {backend}.'
    }


def probar_conexion_correo():
    if _usa_brevo_api():
        return True
    connection = get_connection()
    try:
        opened = connection.open()
        return opened is not False
    finally:
        connection.close()


def registrar_reporte(datos, estado, destinatario='', enviado_por=None, asunto='', mensaje_error=''):
    return ReporteAcudiente.objects.create(
        estudiante=datos['estudiante'],
        periodo=datos['periodo'],
        fecha_inicio=datos['fecha_inicio'],
        fecha_fin=datos['fecha_fin'],
        destinatario=destinatario or None,
        estado=estado,
        enviado_por=enviado_por,
        asunto=asunto,
        mensaje_error=mensaje_error or None,
    )


def enviar_reporte_estudiante_por_correo(datos, enviado_por=None):
    estudiante = datos['estudiante']
    destinatario = estudiante.correo_acudiente
    if not destinatario:
        raise ValueError('El estudiante no tiene correo de acudiente registrado.')

    estado_correo = obtener_estado_correo()
    if not estado_correo['configurado']:
        raise ValueError(estado_correo['mensaje'])

    contenido = generar_excel_reporte_estudiante(datos)
    nombre_archivo = f"reporte_{datos['periodo']}_{estudiante.codigo}.xlsx"

    asunto = f"Reporte {datos['periodo_label'].lower()} de {estudiante.nombres} {estudiante.apellidos}"
    contexto = {
        **datos,
        'destinatario_nombre': estudiante.acudiente,
        'seguimiento_destacado': datos['seguimiento_diario'][:7] if datos['periodo'] == 'semanal' else datos['seguimiento_diario'][:10],
    }
    cuerpo_texto = render_to_string('academico/email/reporte_estudiante.txt', contexto)
    cuerpo_html = render_to_string('academico/email/reporte_estudiante.html', contexto)

    if _usa_brevo_api():
        _enviar_correo_brevo_api(
            asunto=asunto,
            cuerpo_texto=cuerpo_texto,
            cuerpo_html=cuerpo_html,
            destinatario=destinatario,
            nombre_destinatario=estudiante.acudiente,
            nombre_archivo=nombre_archivo,
            contenido=contenido,
        )
    else:
        correo = EmailMultiAlternatives(
            asunto,
            cuerpo_texto,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[destinatario]
        )
        correo.attach_alternative(cuerpo_html, 'text/html')
        correo.attach(nombre_archivo, contenido, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        correo.send(fail_silently=False)
    registrar_reporte(datos, 'ENVIADO', destinatario=destinatario, enviado_por=enviado_por, asunto=asunto)
    return destinatario


def enviar_reportes_estudiantes(estudiantes, periodo, enviado_por=None, fecha_inicio=None, fecha_fin=None):
    enviados = []
    errores = []

    for estudiante in estudiantes:
        try:
            datos = construir_datos_reporte_estudiante(estudiante, periodo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
            destinatario = enviar_reporte_estudiante_por_correo(datos, enviado_por=enviado_por)
            enviados.append((estudiante, destinatario))
        except Exception as exc:
            datos = construir_datos_reporte_estudiante(estudiante, periodo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
            registrar_reporte(
                datos,
                'ERROR',
                destinatario=estudiante.correo_acudiente or '',
                enviado_por=enviado_por,
                asunto=f"Reporte {periodo} de {estudiante.nombres} {estudiante.apellidos}",
                mensaje_error=str(exc),
            )
            errores.append((estudiante, str(exc)))

    return enviados, errores
