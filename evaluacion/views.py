from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from academico.models import Estudiante, PeriodoAcademico
from alertas.utils import evaluar_alertas_academicas
from usuarios.auditoria import registrar_auditoria_cambio
from usuarios.decorators import role_required
from usuarios.notificaciones import crear_notificaciones_docentes
from usuarios.permissions import (
    cargas_visibles_para,
    filtrar_estudiantes_visibles,
    grupos_visibles_para,
    puede_gestionar_docencia,
    no_es_estudiante,
)
from .forms import ActividadEvaluativaForm
from .models import ActividadEvaluativa, Calificacion
from .utils import calcular_promedio_dimensionado


def _url_lista_actividades_filtrada(carga_id=None, periodo_id=None):
    params = {}
    if carga_id:
        params['carga'] = carga_id
    if periodo_id:
        params['periodo'] = periodo_id
    base_url = reverse('lista_actividades')
    return f'{base_url}?{urlencode(params)}' if params else base_url


def _preparar_hoja_excel(hoja, titulo, encabezados):
    hoja.title = titulo
    hoja.append(encabezados)

    fill = PatternFill('solid', fgColor='1F2937')
    for celda in hoja[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = fill


def _ajustar_columnas(hoja):
    for columna in hoja.columns:
        longitud = max(len(str(celda.value or '')) for celda in columna)
        hoja.column_dimensions[get_column_letter(columna[0].column)].width = min(longitud + 2, 45)


def _resumen_dimensiones_actividades(actividades):
    resumen = []
    actividades = list(actividades)

    for clave, etiqueta in ActividadEvaluativa.DIMENSION_CHOICES:
        items = [actividad for actividad in actividades if actividad.dimension == clave]
        objetivo = Decimal(str(ActividadEvaluativa.limite_dimension(clave)))
        peso_dimension = Decimal(str(ActividadEvaluativa.peso_dimension(clave)))
        acumulado = sum((actividad.porcentaje for actividad in items), Decimal('0'))
        disponible = max(Decimal('0'), objetivo - acumulado)
        resumen.append({
            'clave': clave,
            'etiqueta': etiqueta,
            'objetivo': objetivo,
            'peso_dimension': peso_dimension,
            'acumulado': acumulado,
            'disponible': disponible,
            'actividades': items,
            'cantidad': len(items),
            'completa': acumulado == objetivo,
            'avance_periodo': ((acumulado / objetivo) * peso_dimension) if objetivo else Decimal('0'),
        })

    return resumen


def _datos_resumen_calificaciones(request):
    grupo_id = request.GET.get('grupo', '').strip()
    carga_id = request.GET.get('carga', '').strip()
    periodo_id = request.GET.get('periodo', '').strip()
    busqueda = request.GET.get('busqueda', '').strip()

    cargas_visibles = cargas_visibles_para(request.user)
    estudiantes = filtrar_estudiantes_visibles(
        Estudiante.objects.filter(activo=True),
        request.user
    ).select_related('grupo', 'grupo__grado')
    actividades = ActividadEvaluativa.objects.filter(activa=True).select_related(
        'carga_academica',
        'carga_academica__asignatura',
        'carga_academica__grupo',
        'periodo',
    ).filter(carga_academica__in=cargas_visibles)

    if grupo_id:
        estudiantes = estudiantes.filter(grupo_id=grupo_id)
        actividades = actividades.filter(carga_academica__grupo_id=grupo_id)

    if carga_id:
        actividades = actividades.filter(carga_academica_id=carga_id)
        carga = get_object_or_404(cargas_visibles, pk=carga_id)
        estudiantes = estudiantes.filter(grupo=carga.grupo)

    if periodo_id:
        actividades = actividades.filter(periodo_id=periodo_id)

    if busqueda:
        estudiantes = estudiantes.filter(
            Q(codigo__icontains=busqueda) |
            Q(documento__icontains=busqueda) |
            Q(nombres__icontains=busqueda) |
            Q(apellidos__icontains=busqueda)
        )

    actividades = list(actividades.order_by('fecha', 'nombre'))
    actividad_ids = [actividad.id for actividad in actividades]

    calificaciones = Calificacion.objects.filter(
        actividad_id__in=actividad_ids,
        estudiante__in=estudiantes,
    ).select_related('actividad')

    calificaciones_por_estudiante = {}
    for calificacion in calificaciones:
        calificaciones_por_estudiante.setdefault(calificacion.estudiante_id, {})[calificacion.actividad_id] = calificacion

    resumen = []
    for estudiante in estudiantes.order_by('apellidos', 'nombres'):
        notas = calificaciones_por_estudiante.get(estudiante.id, {})
        promedio, resumen_dimensiones = calcular_promedio_dimensionado(
            [calificacion for calificacion in notas.values() if calificacion]
        )

        resumen.append({
            'estudiante': estudiante,
            'notas': [notas.get(actividad.id) for actividad in actividades],
            'promedio': promedio,
            'tiene_promedio': promedio is not None,
            'resumen_dimensiones': resumen_dimensiones,
        })

    detalle_estudiante = None
    if getattr(request.user, 'rol', '') == 'EST' and resumen:
        fila_estudiante = resumen[0]
        estudiante = fila_estudiante['estudiante']
        notas_estudiante = calificaciones_por_estudiante.get(estudiante.id, {})
        materias = []
        materias_por_carga = {}

        for actividad in actividades:
            carga = actividad.carga_academica
            materia = materias_por_carga.get(carga.id)
            if not materia:
                materia = {
                    'carga_id': carga.id,
                    'asignatura': carga.asignatura.nombre,
                    'docente': carga.docente.get_full_name() or carga.docente.username,
                    'grupo': str(carga.grupo),
                    'actividades': [],
                    'calificaciones': [],
                    'promedio': None,
                    'tiene_promedio': False,
                    'resumen_dimensiones': {},
                    'notas_bajas': 0,
                }
                materias_por_carga[carga.id] = materia
                materias.append(materia)

            calificacion = notas_estudiante.get(actividad.id)
            if calificacion:
                materia['calificaciones'].append(calificacion)
                if calificacion.nota < Decimal('3.0'):
                    materia['notas_bajas'] += 1

            materia['actividades'].append({
                'nombre': actividad.nombre,
                'dimension': actividad.get_dimension_display(),
                'porcentaje': actividad.porcentaje,
                'fecha': actividad.fecha,
                'calificacion': calificacion,
                'nota': calificacion.nota if calificacion else None,
                'observacion': calificacion.observacion if calificacion else '',
            })

        for materia in materias:
            promedio_materia, resumen_dimensiones_materia = calcular_promedio_dimensionado(materia['calificaciones'])
            materia['promedio'] = promedio_materia
            materia['tiene_promedio'] = promedio_materia is not None
            materia['resumen_dimensiones'] = resumen_dimensiones_materia
            materia.pop('calificaciones', None)

        detalle_estudiante = {
            'estudiante': estudiante,
            'promedio_general': fila_estudiante['promedio'],
            'tiene_promedio': fila_estudiante['tiene_promedio'],
            'materias': materias,
            'total_materias': len(materias),
            'total_actividades': len(actividades),
        }

    return {
        'resumen': resumen,
        'actividades': actividades,
        'grupos': grupos_visibles_para(request.user),
        'cargas': cargas_visibles,
        'periodos': PeriodoAcademico.objects.select_related('anio_lectivo').filter(activo=True),
        'grupo_id': grupo_id,
        'carga_id': carga_id,
        'periodo_id': periodo_id,
        'busqueda': busqueda,
        'detalle_estudiante': detalle_estudiante,
    }


@role_required(puede_gestionar_docencia)
def lista_actividades(request):
    cargas = cargas_visibles_para(request.user).distinct().order_by(
        'grupo__grado__nombre',
        'grupo__nombre',
        'asignatura__nombre',
    )
    carga_id = request.GET.get('carga', '').strip()
    periodo_id = request.GET.get('periodo', '').strip()
    periodos = PeriodoAcademico.objects.none()
    actividades = ActividadEvaluativa.objects.none()
    carga_seleccionada = None
    periodo_seleccionado = None
    total_actividades = 0
    total_porcentaje = Decimal('0')
    total_porcentaje_disponible = Decimal('100.00')
    total_dimensiones_completas = 0

    if carga_id:
        carga_seleccionada = get_object_or_404(
            cargas.select_related('asignatura', 'grupo', 'grupo__grado', 'anio_lectivo'),
            pk=carga_id
        )
        periodos = PeriodoAcademico.objects.select_related('anio_lectivo').filter(
            anio_lectivo=carga_seleccionada.anio_lectivo,
            activo=True,
        ).order_by('numero', 'fecha_inicio')

        if periodo_id:
            periodo_seleccionado = get_object_or_404(periodos, pk=periodo_id)
            actividades = ActividadEvaluativa.objects.select_related(
                'carga_academica',
                'carga_academica__asignatura',
                'carga_academica__grupo',
                'periodo',
            ).filter(
                carga_academica=carga_seleccionada,
                periodo=periodo_seleccionado,
            ).order_by('fecha', 'nombre')
            total_actividades = actividades.count()

    resumen_dimensiones = _resumen_dimensiones_actividades(actividades)
    total_porcentaje = sum(
        (dimension['avance_periodo'] for dimension in resumen_dimensiones),
        Decimal('0'),
    )
    total_porcentaje_disponible = max(Decimal('0'), Decimal('100.00') - total_porcentaje)
    total_dimensiones_completas = sum(1 for dimension in resumen_dimensiones if dimension['completa'])

    return render(request, 'evaluacion/lista_actividades.html', {
        'actividades': actividades,
        'resumen_dimensiones': resumen_dimensiones,
        'cargas': cargas,
        'periodos': periodos,
        'carga_id': carga_id,
        'periodo_id': periodo_id,
        'carga_seleccionada': carga_seleccionada,
        'periodo_seleccionado': periodo_seleccionado,
        'total_actividades': total_actividades,
        'total_porcentaje': total_porcentaje,
        'total_porcentaje_disponible': total_porcentaje_disponible,
        'total_dimensiones_completas': total_dimensiones_completas,
    })


@role_required(puede_gestionar_docencia)
def crear_actividad(request):
    cargas = cargas_visibles_para(request.user).distinct().order_by(
        'grupo__grado__nombre',
        'grupo__nombre',
        'asignatura__nombre',
    )
    carga_inicial = request.GET.get('carga', '').strip()
    periodo_inicial = request.GET.get('periodo', '').strip()
    dimension_inicial = request.GET.get('dimension', '').strip()

    if request.method == 'POST':
        form = ActividadEvaluativaForm(request.POST, cargas=cargas)
        if form.is_valid():
            actividad = form.save()
            messages.success(request, 'Actividad evaluativa creada correctamente.')
            return redirect(
                _url_lista_actividades_filtrada(
                    carga_id=actividad.carga_academica_id,
                    periodo_id=actividad.periodo_id,
                )
            )
    else:
        form = ActividadEvaluativaForm(
            cargas=cargas,
            initial={
                'carga_academica': carga_inicial or None,
                'periodo': periodo_inicial or None,
                'dimension': dimension_inicial or ActividadEvaluativa.DIMENSION_ACTIVIDADES,
            }
        )

    return render(request, 'evaluacion/actividad_form.html', {
        'form': form,
        'titulo': 'Nueva actividad evaluativa',
        'boton': 'Crear actividad',
        'volver_url': _url_lista_actividades_filtrada(carga_inicial, periodo_inicial),
        'dimension_limites': ActividadEvaluativa.DIMENSION_CHOICES,
    })


@role_required(puede_gestionar_docencia)
def editar_actividad(request, pk):
    actividad = get_object_or_404(ActividadEvaluativa, pk=pk, carga_academica__in=cargas_visibles_para(request.user))
    cargas = cargas_visibles_para(request.user).distinct().order_by(
        'grupo__grado__nombre',
        'grupo__nombre',
        'asignatura__nombre',
    )

    if request.method == 'POST':
        form = ActividadEvaluativaForm(request.POST, instance=actividad, cargas=cargas)
        if form.is_valid():
            actividad = form.save()
            messages.success(request, 'Actividad evaluativa actualizada correctamente.')
            return redirect(
                _url_lista_actividades_filtrada(
                    carga_id=actividad.carga_academica_id,
                    periodo_id=actividad.periodo_id,
                )
            )
    else:
        form = ActividadEvaluativaForm(instance=actividad, cargas=cargas)

    return render(request, 'evaluacion/actividad_form.html', {
        'form': form,
        'actividad': actividad,
        'titulo': 'Editar actividad evaluativa',
        'boton': 'Guardar cambios',
        'volver_url': _url_lista_actividades_filtrada(
            actividad.carga_academica_id,
            actividad.periodo_id,
        ),
        'dimension_limites': ActividadEvaluativa.DIMENSION_CHOICES,
    })


@role_required(puede_gestionar_docencia)
def eliminar_actividad(request, pk):
    actividad = get_object_or_404(ActividadEvaluativa, pk=pk, carga_academica__in=cargas_visibles_para(request.user))

    if request.method == 'POST':
        redirect_url = _url_lista_actividades_filtrada(
            actividad.carga_academica_id,
            actividad.periodo_id,
        )
        actividad.delete()
        messages.success(request, 'Actividad evaluativa eliminada correctamente.')
        return redirect(redirect_url)

    return render(request, 'evaluacion/confirmar_eliminar_actividad.html', {
        'actividad': actividad,
    })


@role_required(puede_gestionar_docencia)
def registrar_calificaciones(request, pk):
    actividad = get_object_or_404(
        ActividadEvaluativa.objects.select_related(
            'carga_academica',
            'carga_academica__grupo',
            'carga_academica__asignatura',
            'periodo',
        ),
        pk=pk,
        carga_academica__in=cargas_visibles_para(request.user)
    )

    estudiantes = Estudiante.objects.filter(
        grupo=actividad.carga_academica.grupo,
        activo=True,
    ).order_by('apellidos', 'nombres')

    calificaciones = {
        calificacion.estudiante_id: calificacion
        for calificacion in Calificacion.objects.filter(actividad=actividad)
    }
    usar_observacion = any(clave.startswith('observacion_') for clave in request.POST.keys())

    errores = []

    if request.method == 'POST':
        valores_enviados = {}
        cambios_realizados = []
        detalle_cambios = []

        with transaction.atomic():
            for estudiante in estudiantes:
                nota_raw = request.POST.get(f'nota_{estudiante.id}', '').strip().replace(',', '.')
                calificacion_previa = calificaciones.get(estudiante.id)
                observacion = request.POST.get(f'observacion_{estudiante.id}', '').strip() if usar_observacion else (
                    calificacion_previa.observacion if calificacion_previa else ''
                )
                valores_enviados[estudiante.id] = {
                    'nota': nota_raw,
                    'observacion': observacion,
                }

                if nota_raw == '':
                    if calificacion_previa:
                        cambios_realizados.append(
                            f'{estudiante.apellidos} {estudiante.nombres}: {calificacion_previa.nota} -> sin nota'
                        )
                        detalle_cambios.append({
                            'estudiante': f'{estudiante.apellidos} {estudiante.nombres}',
                            'valor_anterior': str(calificacion_previa.nota),
                            'valor_nuevo': 'Sin nota',
                            'observacion_anterior': calificacion_previa.observacion or '',
                            'observacion_nueva': '',
                            'actividad': actividad.nombre,
                        })
                        registrar_auditoria_cambio(
                            actor=request.user,
                            tipo='CALIFICACION',
                            accion='ELIMINACION',
                            modulo='calificaciones',
                            titulo=f'Eliminacion de nota en {actividad.carga_academica.asignatura.nombre}',
                            descripcion=(
                                f'{request.user.get_full_name() or request.user.username} elimino la nota de '
                                f'{estudiante.apellidos} {estudiante.nombres} en {actividad.nombre} '
                                f'({actividad.carga_academica.grupo}).'
                            ),
                            estudiante=estudiante,
                            grupo=str(actividad.carga_academica.grupo),
                            asignatura=actividad.carga_academica.asignatura.nombre,
                            fecha_referencia=actividad.fecha,
                            valor_anterior=calificacion_previa.nota,
                            valor_nuevo='Sin nota',
                            referencia_url=f'/evaluacion/actividad/{actividad.pk}/calificaciones/',
                            datos_extra={
                                'actividad': actividad.nombre,
                                'observacion_anterior': calificacion_previa.observacion or '',
                                'observacion_nueva': '',
                            },
                        )
                    Calificacion.objects.filter(actividad=actividad, estudiante=estudiante).delete()
                    evaluar_alertas_academicas(estudiante)
                    continue

                try:
                    nota = Decimal(nota_raw)
                except InvalidOperation:
                    errores.append(f'{estudiante}: la nota ingresada no es valida.')
                    continue

                if nota < Decimal('0') or nota > Decimal('5'):
                    errores.append(f'{estudiante}: la nota debe estar entre 0.00 y 5.00.')
                    continue

                if calificacion_previa and (
                    calificacion_previa.nota != nota or
                    (calificacion_previa.observacion or '') != observacion
                ):
                    cambios_realizados.append(
                        f'{estudiante.apellidos} {estudiante.nombres}: {calificacion_previa.nota} -> {nota}'
                    )
                    detalle_cambios.append({
                        'estudiante': f'{estudiante.apellidos} {estudiante.nombres}',
                        'valor_anterior': str(calificacion_previa.nota),
                        'valor_nuevo': str(nota),
                        'observacion_anterior': calificacion_previa.observacion or '',
                        'observacion_nueva': observacion,
                        'actividad': actividad.nombre,
                    })

                accion = 'CREACION' if calificacion_previa is None else 'EDICION'

                Calificacion.objects.update_or_create(
                    actividad=actividad,
                    estudiante=estudiante,
                    defaults={
                        'nota': nota,
                        'observacion': observacion,
                    }
                )
                if (
                    calificacion_previa is None or
                    calificacion_previa.nota != nota or
                    (calificacion_previa.observacion or '') != observacion
                ):
                    registrar_auditoria_cambio(
                        actor=request.user,
                        tipo='CALIFICACION',
                        accion=accion,
                        modulo='calificaciones',
                        titulo=f'Actualizacion de nota en {actividad.carga_academica.asignatura.nombre}',
                        descripcion=(
                            f'{request.user.get_full_name() or request.user.username} actualizo la calificacion de '
                            f'{estudiante.apellidos} {estudiante.nombres} en {actividad.nombre} '
                            f'({actividad.carga_academica.grupo}).'
                        ),
                        estudiante=estudiante,
                        grupo=str(actividad.carga_academica.grupo),
                        asignatura=actividad.carga_academica.asignatura.nombre,
                        fecha_referencia=actividad.fecha,
                        valor_anterior=calificacion_previa.nota if calificacion_previa else 'Sin nota',
                        valor_nuevo=nota,
                        referencia_url=f'/evaluacion/actividad/{actividad.pk}/calificaciones/',
                        datos_extra={
                            'actividad': actividad.nombre,
                            'observacion_anterior': (calificacion_previa.observacion or '') if calificacion_previa else '',
                            'observacion_nueva': observacion,
                            'porcentaje': str(actividad.porcentaje),
                        },
                    )
                evaluar_alertas_academicas(estudiante)

            if errores:
                transaction.set_rollback(True)
            else:
                if cambios_realizados:
                    resumen = '; '.join(cambios_realizados[:4])
                    if len(cambios_realizados) > 4:
                        resumen += f'; y {len(cambios_realizados) - 4} cambio(s) mas'
                    crear_notificaciones_docentes(
                        request.user,
                        'CALIFICACION',
                        f'Modificacion de notas en {actividad.carga_academica.asignatura.nombre} - {actividad.carga_academica.grupo}',
                        (
                            f'Se modificaron {len(cambios_realizados)} calificaciones en "{actividad.nombre}" '
                            f'por {request.user.get_full_name() or request.user.username}. {resumen}.'
                        ),
                        url=f'/evaluacion/actividad/{actividad.pk}/calificaciones/',
                        detalle_cambios=detalle_cambios,
                        metadata={
                            'asignatura': actividad.carga_academica.asignatura.nombre,
                            'grupo': str(actividad.carga_academica.grupo),
                            'fecha': str(actividad.fecha),
                            'actividad': actividad.nombre,
                        },
                    )
                messages.success(request, 'Calificaciones guardadas correctamente.')
                return redirect('registrar_calificaciones', pk=actividad.pk)

        for estudiante in estudiantes:
            enviado = valores_enviados.get(estudiante.id, {})
            estudiante.nota_guardada = enviado.get('nota', '')
            estudiante.observacion_guardada = enviado.get('observacion', '')
    else:
        for estudiante in estudiantes:
            calificacion = calificaciones.get(estudiante.id)
            estudiante.nota_guardada = calificacion.nota if calificacion else ''
            estudiante.observacion_guardada = calificacion.observacion if calificacion else ''

    return render(request, 'evaluacion/registrar_calificaciones.html', {
        'actividad': actividad,
        'estudiantes': estudiantes,
        'errores': errores,
    })


@login_required
@role_required(no_es_estudiante)
def resumen_calificaciones(request):
    return render(
        request,
        'evaluacion/resumen_calificaciones.html',
        _datos_resumen_calificaciones(request)
    )


@login_required
@role_required(no_es_estudiante)
def exportar_resumen_calificaciones_excel(request):
    datos = _datos_resumen_calificaciones(request)

    libro = Workbook()
    hoja = libro.active
    encabezados = ['Codigo', 'Documento', 'Estudiante', 'Grado', 'Grupo']
    encabezados += [
        f'{actividad.carga_academica.asignatura.nombre} - {actividad.nombre} ({actividad.porcentaje}%)'
        for actividad in datos['actividades']
    ]
    encabezados.append('Promedio ponderado')
    _preparar_hoja_excel(hoja, 'Calificaciones', encabezados)

    for fila in datos['resumen']:
        estudiante = fila['estudiante']
        hoja.append([
            estudiante.codigo,
            estudiante.documento,
            f'{estudiante.apellidos} {estudiante.nombres}',
            estudiante.grupo.grado.nombre,
            estudiante.grupo.nombre,
            *[calificacion.nota if calificacion else '' for calificacion in fila['notas']],
            fila['promedio'] if fila['tiene_promedio'] else '',
        ])

    _ajustar_columnas(hoja)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="resumen_calificaciones.xlsx"'
    libro.save(response)
    return response
